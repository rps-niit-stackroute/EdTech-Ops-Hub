import openpyxl
from openpyxl.utils import get_column_letter
wb = openpyxl.load_workbook("/app/samples/out_test.xlsx")
ws = wb["Consolidated Report"]
print("Consolidated appended cols 66/67 (BN/BO):")
for r in [1,2,3,4]:
    for c in [64,65,66,67]:
        cell = ws.cell(row=r,column=c)
        f=cell.fill
        fg = (f.fgColor.rgb if f.fgColor.type=='rgb' else 'theme'+str(f.fgColor.theme)) if f.patternType else None
        b = cell.border
        bsides = "".join(s for s,side in [("L",b.left),("R",b.right),("T",b.top),("B",b.bottom)] if side and side.style)
        print(f"  {get_column_letter(c)}{r} val={cell.value!r} pat={f.patternType} fg={fg} border={bsides} font={cell.font.name},{cell.font.size}")
print("merged ranges near end:", [str(m) for m in ws.merged_cells.ranges if m.min_col>=64])

ws3 = wb["Login"]
print("\nLogin appended group cols 96-99:")
for r in [1,3,4]:
    for c in [96,97,98,99]:
        cell = ws3.cell(row=r,column=c)
        f=cell.fill
        fg = (f.fgColor.rgb if f.fgColor.type=='rgb' else 'theme'+str(f.fgColor.theme)) if f.patternType else None
        print(f"  {get_column_letter(c)}{r} val={cell.value!r} pat={f.patternType} fg={fg}")
print("login merged near end:", [str(m) for m in ws3.merged_cells.ranges if m.min_col>=95])
