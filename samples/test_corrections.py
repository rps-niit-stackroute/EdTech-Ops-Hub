import sys; sys.path.insert(0,'/app/backend')
import openpyxl, io
from attendance_processor import parse_teams_export, process_attendance

teams = """1. Summary
Meeting title\tCNA Level 1
Meeting Start Time\t5/22/26, 9:30:00 AM
Meeting Duration\t2h 30m
2. Participants
Name\tFirst Join\tLast Leave\tIn-Meeting Duration\tEmail\tRole
Rajan, Seshadri (Cognizant)\t5/22/26, 9:30:00 AM\t5/22/26, 12:00:00 PM\t2h 28m 0s\ta@x.com\tAttendee
Gupta, Mrudul (Cognizant)\t5/22/26, 9:35:00 AM\t5/22/26, 12:00:00 PM\t1h 5m 0s\tb@x.com\tAttendee
P, Naveen (Cognizant)\t5/22/26, 9:30:00 AM\t5/22/26, 11:00:00 AM\t30m 0s\tc@x.com\tAttendee
3. In-Meeting Activities
""".encode("utf-16")

parsed = parse_teams_export(teams, "t.csv")
print("detected date:", parsed["session_date"], "| max_minutes:", parsed["max_minutes"])

tb = open("/app/samples/tracker.xlsx","rb").read()
out, name, info = process_attendance(tb, "NIIT_CNA_20thMay.xlsx", teams, "t.csv", "Retake", "2026-05-22", 50)
wb = openpyxl.load_workbook(io.BytesIO(out))
ws = wb["Consolidated Report"]
mc = ws.max_column
print("\nAttentiveness data (BO col), expect Seshadri=148/148=100%, Mrudul=65/148, Naveen=30/148:")
for r in range(3, 8):
    nm = ws.cell(row=r, column=2).value
    att = ws.cell(row=r, column=mc)
    atd = ws.cell(row=r, column=mc-1)
    print(f"  {nm}: attend={atd.value} attentiveness={att.value} fmt={att.number_format}")

# SOW data test
print("\n--- SOW dates + excel ---")
import asyncio
from server import build_sow_data
from sow_export import build_sow_excel
import datetime
g, grand, ml = asyncio.get_event_loop().run_until_complete(
    build_sow_data(str(datetime.date.today().month), datetime.date.today().year, "", ""))
for grp in g[:1]:
    for row in grp["rows"]:
        print("  row dates:", row["dates"], "| sessions:", row["sessions_conducted"])
xls = build_sow_excel(g, ml)
wb2 = openpyxl.load_workbook(io.BytesIO(xls))
ws2 = wb2.active
print("  SOW header:", [ws2.cell(row=1,column=c).value for c in range(1,10)])
print("  row2:", [ws2.cell(row=2,column=c).value for c in range(1,10)])
print("  last row:", [ws2.cell(row=ws2.max_row,column=c).value for c in range(1,10)])
