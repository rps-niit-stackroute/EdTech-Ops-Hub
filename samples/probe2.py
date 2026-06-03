import openpyxl
from openpyxl.utils import get_column_letter

def probe(path, sheet=None, rstart=1, rend=20, cstart=1, cend=12):
    wb = openpyxl.load_workbook(path, data_only=False)
    print("="*70, "\nFILE:", path, "SHEETS:", wb.sheetnames)
    sheets = [sheet] if sheet else wb.sheetnames
    for sn in sheets:
        ws = wb[sn]
        print("-"*50, f"\nSHEET '{sn}' max_row={ws.max_row} max_col={ws.max_column}")
        print("MERGED:", [str(m) for m in list(ws.merged_cells.ranges)[:20]])
        for r in range(rstart, min(rend, ws.max_row)+1):
            cells=[]
            for c in range(cstart, min(cend, ws.max_column)+1):
                cell=ws.cell(row=r,column=c); v=cell.value
                if v is not None:
                    f=cell.fill
                    fg=None
                    if f and f.patternType:
                        fg = f.fgColor.rgb if f.fgColor.type=='rgb' else ('theme'+str(f.fgColor.theme))
                    cells.append(f"{get_column_letter(c)}{r}={v!r}{('<'+str(fg)+'>') if fg else ''}")
            if cells: print("  "," | ".join(cells))

# SOW sample
probe("/app/samples/sow.xlsx", rend=30, cend=10)
# Lowes (a single-program calendar tab maybe)
probe("/app/samples/lowes.xlsx", rend=15, cend=12)
# bottom of overall attendance
print("\n\n##### BOTTOM Overall Attendance #####")
probe("/app/samples/tracker.xlsx", sheet='Overall Attendance %', rstart=10, rend=26, cend=6)
