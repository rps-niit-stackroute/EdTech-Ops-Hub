import sys
sys.path.insert(0, "/app/backend")
import openpyxl, io

# Build a synthetic Teams export CSV (tab-separated, like real Teams)
teams = """1. Summary
Meeting title\tCNA Level 1
Meeting Start Time\t5/22/26, 9:30:00 AM
Meeting End Time\t5/22/26, 12:00:00 PM
Meeting Duration\t2h 30m
Number of Participants\t5
2. Participants
Name\tFirst Join\tLast Leave\tIn-Meeting Duration\tEmail\tRole
Rajan, Seshadri (Cognizant)\t5/22/26, 9:30:00 AM\t5/22/26, 12:00:00 PM\t2h 28m 0s\ta@x.com\tAttendee
Gupta, Mrudul (Cognizant)\t5/22/26, 9:35:00 AM\t5/22/26, 12:00:00 PM\t1h 5m 0s\tb@x.com\tAttendee
P, Naveen (Cognizant)\t5/22/26, 9:30:00 AM\t5/22/26, 11:00:00 AM\t30m 0s\tc@x.com\tAttendee
Saleem, Abdul Baseer (Cognizant)\t5/22/26, 9:30:00 AM\t5/22/26, 12:00:00 PM\t2h 10m 0s\td@x.com\tAttendee
Unknown, Random Guest (Acme)\t5/22/26, 9:30:00 AM\t5/22/26, 12:00:00 PM\t2h 0m 0s\te@x.com\tAttendee
3. In-Meeting Activities
Name\tJoin Time\tLeave Time
"""
teams_bytes = teams.encode("utf-16")  # Teams uses UTF-16

from attendance_processor import process_attendance, parse_teams_export
parsed = parse_teams_export(teams_bytes, "teams.csv")
print("session_minutes:", parsed["session_minutes"])
print("participants:", [(p["name"], round(p["minutes"],1)) for p in parsed["participants"]])

tracker_bytes = open("/app/samples/tracker.xlsx","rb").read()
out, name, info = process_attendance(tracker_bytes, "NIIT_SR CTS_ CNA Level-1 20th May (2).xlsx",
                                     teams_bytes, "teams.csv", "Final Introspect Retake", "2026-05-22", 50)
print("OUTPUT FILENAME:", name)
print("INFO:", info)
# verify output opens and new columns added
wb = openpyxl.load_workbook(io.BytesIO(out))
ws = wb["Consolidated Report"]
print("Consolidated max_col now:", ws.max_column, "last header:", ws.cell(row=1,column=ws.max_column-1).value)
ws2 = wb["Overall Attendance %"]
print("Overall max_col now:", ws2.max_column, "header:", ws2.cell(row=1,column=ws2.max_column).value, "sub:", ws2.cell(row=2,column=ws2.max_column).value)
ws3 = wb["Login"]
print("Login max_col now:", ws3.max_column)
open("/app/samples/out_test.xlsx","wb").write(out)
print("OK saved out_test.xlsx")

# schedule parser test with Lowes
from schedule_parser import parse_schedule
lo = parse_schedule(open("/app/samples/lowes.xlsx","rb").read(), "lowes.xlsx")
print("\nLowes parsed sessions:", len(lo))
for s in lo[:3]: print(" ", s)
