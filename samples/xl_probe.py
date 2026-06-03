import openpyxl
from openpyxl.utils import get_column_letter

def fill_hex(cell):
    f = cell.fill
    if f and f.fgColor and f.patternType:
        return f.patternType, (f.fgColor.rgb if f.fgColor.type == 'rgb' else f.fgColor.theme)
    return None

def inspect(path, max_rows=12, max_cols=30):
    print("="*80)
    print("FILE:", path)
    wb = openpyxl.load_workbook(path, data_only=False)
    print("SHEETS:", wb.sheetnames)
    for ws in wb.worksheets:
        print("-"*70)
        print(f"SHEET: '{ws.title}'  dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
        print("MERGED:", [str(m) for m in list(ws.merged_cells.ranges)[:30]])
        for r in range(1, min(max_rows, ws.max_row)+1):
            rowvals = []
            for c in range(1, min(max_cols, ws.max_column)+1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v is not None:
                    fh = fill_hex(cell)
                    rowvals.append(f"{get_column_letter(c)}{r}={v!r}{('['+str(fh)+']') if fh else ''}")
            if rowvals:
                print("  ", " | ".join(rowvals))

for f in ["tracker.xlsx","calendar.xlsx","sow.xlsx","lowes.xlsx"]:
    try:
        inspect("/app/samples/"+f)
    except Exception as e:
        print("ERR", f, e)
