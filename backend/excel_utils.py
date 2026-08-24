"""Shared Excel/date/time parsing helpers."""
import re
import os
from copy import copy
from datetime import datetime, date

from openpyxl.utils import get_column_letter

MONTHS_FULL = ["january", "february", "march", "april", "may", "june", "july",
               "august", "september", "october", "november", "december"]
MONTHS_ABBR = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep",
               "oct", "nov", "dec"]


def ordinal(n):
    if 10 <= n % 100 <= 20:
        suf = "th"
    else:
        suf = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return suf


def parse_time_to_24h(value):
    """Parse '10:00 AM', '10:00', '6PM', '6 PM', '12:00 PM' -> 'HH:MM' (24h). Returns None on fail."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    s = str(value).strip().upper()
    if not s:
        return None
    s = s.replace("(IST)", "").strip()
    m = re.search(r'(\d{1,2})\s*[:\.]?\s*(\d{2})?\s*(AM|PM)?', s)
    if not m:
        return None
    hh = int(m.group(1))
    mm = int(m.group(2)) if m.group(2) else 0
    ap = m.group(3)
    if ap == "AM" and hh == 12:
        hh = 0
    elif ap == "PM" and hh != 12:
        hh += 12
    if hh > 23 or mm > 59:
        return None
    return f"{hh:02d}:{mm:02d}"


def time_to_minutes(hhmm):
    if not hhmm:
        return None
    try:
        h, m = str(hhmm).split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return None


_DATE_FORMATS = ["%Y-%m-%d", "%d-%b-%Y", "%d-%B-%Y", "%d/%m/%Y", "%m/%d/%Y",
                "%d-%m-%Y", "%d %b %Y", "%d %B %Y", "%Y/%m/%d", "%d/%m/%y",
                "%m/%d/%y", "%d-%b-%y"]


def _parse_ordinal_date(s):
    """Fallback for 'day ordinal month year' text, e.g. '20th May 2026'."""
    m = re.search(r'(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s*(\d{4})?', s)
    if not m:
        return None
    day = int(m.group(1))
    mon = m.group(2).lower()[:3]
    if mon not in MONTHS_ABBR:
        return None
    yr = int(m.group(3)) if m.group(3) else datetime.now().year
    try:
        return date(yr, MONTHS_ABBR.index(mon) + 1, day).strftime("%Y-%m-%d")
    except Exception:
        return None


def parse_date_flexible(value):
    """Return ISO 'YYYY-MM-DD' from many formats / datetime, else None."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    for f in _DATE_FORMATS:
        try:
            return datetime.strptime(s, f).strftime("%Y-%m-%d")
        except Exception:
            continue
    return _parse_ordinal_date(s)


def parse_duration_to_minutes(value):
    """Parse '1h 7m 35s', '55m 13s', '3h', '01:07:35', '2h 59m 52s' -> float minutes."""
    if value is None:
        return 0.0
    s = str(value).strip()
    if not s:
        return 0.0
    # HH:MM:SS or H:MM:SS
    cm = re.fullmatch(r'(\d{1,3}):(\d{1,2}):(\d{1,2})', s)
    if cm:
        h, m, sec = int(cm.group(1)), int(cm.group(2)), int(cm.group(3))
        return h * 60 + m + sec / 60.0
    # token style — the hour/minute/second count can itself be a decimal (e.g. "1.5h")
    total = 0.0
    found = False
    hm = re.search(r'(\d{1,4}(?:\.\d{1,4})?)\s*h', s, re.I)
    mm = re.search(r'(\d{1,4}(?:\.\d{1,4})?)\s*m(?!s)', s, re.I)
    sm = re.search(r'(\d{1,4}(?:\.\d{1,4})?)\s*s', s, re.I)
    if hm:
        total += float(hm.group(1)) * 60
        found = True
    if mm:
        total += float(mm.group(1))
        found = True
    if sm:
        total += float(sm.group(1)) / 60.0
        found = True
    if found:
        return total
    # plain number = minutes
    try:
        return float(s)
    except Exception:
        return 0.0


_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def sanitize_excel_value(value):
    """Defuse CSV/Excel formula injection. A string cell value starting with
    =, +, -, or @ is interpreted by Excel as a live formula regardless of how
    it was written into the file — and several values we write come from
    sources we don't control (a Teams participant's own display name, a
    free-typed program/mentor/client name). Prefixing with a single quote
    forces Excel to treat it as literal text instead."""
    if isinstance(value, str) and value.startswith(_FORMULA_TRIGGER_CHARS):
        return "'" + value
    return value


def copy_cell_style(src, dst):
    """Copy visual style (font, fill, border, alignment, number format) src->dst cell."""
    try:
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
    except Exception:
        pass


def normalize_teams_name(raw):
    """'Saleem, Abdul Baseer (Cognizant)' -> 'Abdul Baseer Saleem'. Strips company in parens."""
    if not raw:
        return ""
    s = str(raw).strip()
    s = re.sub(r'\([^)]*\)', '', s).strip()  # strip company
    if "," in s:
        parts = s.split(",", 1)
        last = parts[0].strip()
        first = parts[1].strip()
        s = f"{first} {last}".strip()
    return re.sub(r'\s+', ' ', s).strip()


# Pattern: day + optional ordinal + optional separator + month name (abbr or full),
# with an optional trailing year (numeric or separated) to also update.
_MONTH_NAME_PAT = re.compile(
    r'(\d{1,2})(st|nd|rd|th)?([\s\-_]*)([A-Za-z]{3,9})(?:([\s\-_]+)(\d{2,4}))?'
)
# Pattern: ISO numeric date, e.g. 2026-05-22, 2026_05_22, 2026/05/22
_ISO_DATE_PAT = re.compile(r'(?<!\d)(\d{4})([-_/])(\d{1,2})([-_/])(\d{1,2})(?!\d)')
# Pattern: day-month-year numeric, e.g. 22-05-26, 22_05_2026, 22/05/26
_NUMERIC_DATE_PAT = re.compile(r'(?<!\d)(\d{1,2})([-_/])(\d{1,2})([-_/])(\d{2,4})(?!\d)')


def _month_name_repl(m, nd):
    month_tok = m.group(4)
    mtl = month_tok.lower()
    is_full = mtl in MONTHS_FULL
    is_abbr = mtl in MONTHS_ABBR
    if not (is_full or is_abbr):
        return m.group(0)
    day = nd.day
    had_ord = m.group(2) is not None
    sep = m.group(3)
    year_sep, year_tok = m.group(5), m.group(6)
    new_month = nd.strftime("%B") if is_full else nd.strftime("%b")
    # preserve case style of original token
    if month_tok.isupper():
        new_month = new_month.upper()
    elif month_tok.islower():
        new_month = new_month.lower()
    out = str(day)
    if had_ord:
        out += ordinal(day)
    out += sep + new_month
    if year_tok:
        new_year = nd.strftime("%y") if len(year_tok) <= 2 else nd.strftime("%Y")
        out += year_sep + new_year
    return out


def _iso_date_repl(m, nd):
    _, sep1, m_tok, sep2, d_tok = m.groups()
    new_year = nd.strftime("%Y")
    new_month = f"{nd.month:0{len(m_tok)}d}"
    new_day = f"{nd.day:0{len(d_tok)}d}"
    return f"{new_year}{sep1}{new_month}{sep2}{new_day}"


def _numeric_day_first(d_val, m_val):
    """True/False for day-first vs. month-first, or None if the pair can't be a date."""
    if d_val > 12 and m_val <= 12:
        return True
    if m_val > 12 and d_val <= 12:
        return False
    if d_val > 31 or (d_val > 12 and m_val > 12):
        return None
    return True


def _numeric_date_repl(m, nd):
    d_tok, sep1, m_tok, sep2, y_tok = m.group(1), m.group(2), m.group(3), m.group(4), m.group(5)
    try:
        d_val, m_val = int(d_tok), int(m_tok)
    except Exception:
        return m.group(0)
    day_first = _numeric_day_first(d_val, m_val)
    if day_first is None:
        return m.group(0)  # not actually a date
    new_day = f"{nd.day:0{len(d_tok)}d}"
    new_month = f"{nd.month:0{len(m_tok)}d}"
    new_year = nd.strftime("%y") if len(y_tok) <= 2 else nd.strftime("%Y")
    if day_first:
        return f"{new_day}{sep1}{new_month}{sep2}{new_year}"
    return f"{new_month}{sep1}{new_day}{sep2}{new_year}"


def rename_with_new_date(orig_filename, new_iso_date):
    """Replace the date segment in the filename with new date, keeping the same format.
    e.g. NIIT_CNA_20thMay.xlsx + 2026-05-22 -> NIIT_CNA_22ndMay.xlsx
    e.g. tracker_22-05-26.xlsx + 2026-06-15 -> tracker_15-06-26.xlsx
    """
    try:
        nd = datetime.strptime(new_iso_date, "%Y-%m-%d")
    except Exception:
        return orig_filename
    base, ext = os.path.splitext(orig_filename)
    if not ext:
        ext = ".xlsx"

    for pattern, repl in (
        (_MONTH_NAME_PAT, _month_name_repl),
        (_ISO_DATE_PAT, _iso_date_repl),
        (_NUMERIC_DATE_PAT, _numeric_date_repl),
    ):
        new_base, n = pattern.subn(lambda m: repl(m, nd), base, count=1)
        if n > 0:
            return new_base + ext

    # No recognizable date segment -> append new date
    return base + "_" + nd.strftime("%d-%b-%Y") + ext


def copy_column_width(ws, src_col, dst_col):
    """Copy a column's width from src_col to dst_col (1-indexed column numbers)."""
    src_dim = ws.column_dimensions.get(get_column_letter(src_col))
    if src_dim is not None and src_dim.width:
        ws.column_dimensions[get_column_letter(dst_col)].width = src_dim.width
