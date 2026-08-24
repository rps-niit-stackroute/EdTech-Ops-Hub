"""Attendance tracker append processor — parses Teams export and appends to 3 sheets,
preserving exact format/colors/merges by copying existing cell styles."""
import io
import csv
import re
import difflib
from copy import copy

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_utils import (
    parse_duration_to_minutes, normalize_teams_name, copy_cell_style,
    rename_with_new_date, copy_column_width, sanitize_excel_value,
)
from feedback_processor import parse_feedback_export, append_feedback_sheet


# ----------------------------- Teams export parsing -----------------------------
def _decode_bytes(b):
    for enc in ("utf-16", "utf-8-sig", "utf-8", "latin-1"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return b.decode("utf-8", errors="ignore")


def _rows_from_csv(b):
    text = _decode_bytes(b)
    delim = "\t" if text.count("\t") >= text.count(",") else ","
    rows = []
    for line in text.splitlines():
        rows.append(next(csv.reader([line], delimiter=delim)))
    return rows


def _rows_from_xlsx(b):
    wb = openpyxl.load_workbook(io.BytesIO(b), data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c) for c in r])
    return rows


def _first_duration_after(r, start_idx):
    """Parsed duration of the first non-empty cell after start_idx in this row,
    or None if the row has no non-empty cell past that point."""
    for j in range(start_idx + 1, len(r)):
        if str(r[j]).strip():
            return parse_duration_to_minutes(r[j])
    return None


def _find_session_minutes(rows):
    """Meeting duration from the Teams summary section — the value cell directly
    after whichever cell has the "Meeting Duration" label."""
    for r in rows:
        for i, c in enumerate(r):
            if not (c and re.search(r'meeting\s+duration', str(c), re.I)):
                continue
            minutes = _first_duration_after(r, i)
            if minutes:
                return minutes
    return 0.0


def _map_participant_column(col_map, ci, val):
    if val == "name" and "name" not in col_map:
        col_map["name"] = ci
    elif "first join" in val:
        col_map["first_join"] = ci
    elif "last leave" in val:
        col_map["last_leave"] = ci
    elif "duration" in val:
        col_map["duration"] = ci
    elif "attentiveness" in val or "attention" in val:
        col_map["attentiveness"] = ci
    elif "email" in val or "upn" in val:
        col_map["email"] = ci


def _find_participants_header(rows):
    """(header_idx, col_map) for the participants table, or (None, {}) if not found."""
    for idx, r in enumerate(rows):
        low = [str(c).strip().lower() for c in r]
        if "name" in low and any("duration" in x for x in low):
            col_map = {}
            for ci, val in enumerate(low):
                _map_participant_column(col_map, ci, val)
            return idx, col_map
    return None, {}


def _parse_one_participant_row(r, col_map):
    """Participant dict for this row, or None to skip it (no name, or a repeated header row)."""
    name_raw = str(r[col_map["name"]]).strip() if col_map["name"] < len(r) else ""
    if not name_raw or name_raw.strip().lower() == "name":
        return None
    dur = r[col_map["duration"]] if col_map["duration"] < len(r) else ""
    return {
        "raw": name_raw,
        "name": normalize_teams_name(name_raw),
        "minutes": parse_duration_to_minutes(dur),
        "first_join": str(r[col_map["first_join"]]).strip() if col_map.get("first_join", 99) < len(r) else "",
        "last_leave": str(r[col_map["last_leave"]]).strip() if col_map.get("last_leave", 99) < len(r) else "",
        "email": str(r[col_map["email"]]).strip() if col_map.get("email", 99) < len(r) else "",
    }


def _parse_participant_rows(rows, header_idx, col_map):
    participants = []
    if header_idx is None or "name" not in col_map or "duration" not in col_map:
        return participants
    for r in rows[header_idx + 1:]:
        if not r:
            continue
        # stop at next section marker like "3. In-Meeting Activities"
        first_cell = str(r[0]).strip()
        if re.match(r'^\d+\.\s+\w', first_cell):
            break
        participant = _parse_one_participant_row(r, col_map)
        if participant:
            participants.append(participant)
    return participants


def parse_teams_export(file_bytes, filename):
    """Returns dict: {participants:[{raw,name,minutes,attentiveness,first_join,last_leave}],
                      session_minutes: float, attentiveness_available: bool}"""
    if filename.lower().endswith((".xlsx", ".xls")):
        rows = _rows_from_xlsx(file_bytes)
    else:
        rows = _rows_from_csv(file_bytes)

    session_minutes = _find_session_minutes(rows)
    header_idx, col_map = _find_participants_header(rows)
    participants = _parse_participant_rows(rows, header_idx, col_map)

    if not session_minutes:
        session_minutes = max((p["minutes"] for p in participants), default=0.0)

    # Highest duration among all participants -> denominator for attentiveness
    max_minutes = max((p["minutes"] for p in participants), default=0.0)

    return {
        "participants": participants,
        "session_minutes": session_minutes or 0.0,
        "max_minutes": max_minutes or 0.0,
        "session_date": _detect_session_date(rows, participants),
    }


def _first_nonempty_after(r, start_idx):
    for j in range(start_idx + 1, len(r)):
        if str(r[j]).strip():
            return str(r[j])
    return None


def _meeting_start_candidates(rows):
    candidates = []
    for r in rows:
        for i, c in enumerate(r):
            if not (c and re.search(r'meeting\s+start', str(c), re.I)):
                continue
            val = _first_nonempty_after(r, i)
            if val:
                candidates.append(val)
    return candidates


def _detect_session_date(rows, participants):
    """Detect the session date (ISO) from Teams summary 'Meeting Start Time' or a participant join."""
    candidates = _meeting_start_candidates(rows)
    if participants:
        candidates.append(participants[0].get("first_join", ""))
    for cand in candidates:
        iso = _parse_teams_datetime(cand)
        if iso:
            return iso
    return None


def _parse_teams_datetime(s):
    if not s:
        return None
    first = str(s).split(",")[0].strip()
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            from datetime import datetime as _dt
            return _dt.strptime(first, fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    return None


# ----------------------------- Name matching -----------------------------
def _norm(s):
    return re.sub(r'\s+', ' ', str(s or "").lower()).strip()


def _match_exact(en, p_norm, used):
    for i, (pn, p) in enumerate(p_norm):
        if i not in used and pn == en:
            return p, i
    return None, None


def _match_partial(en_parts, p_norm, used):
    """All enrolled name tokens present in the participant's name."""
    for i, (pn, p) in enumerate(p_norm):
        if i in used:
            continue
        pset = set(pn.split())
        if en_parts and all(t in pset for t in en_parts):
            return p, i
    return None, None


def _match_first_name(en_parts, p_norm, used):
    if not en_parts:
        return None, None
    for i, (pn, p) in enumerate(p_norm):
        if i in used:
            continue
        if en_parts[0] and en_parts[0] in pn.split():
            return p, i
    return None, None


# Below this ratio, two names are treated as unrelated rather than a typo/reorder
# of each other — chosen to catch things like "Rahul Sharma" vs "Rahul Sharme"
# (typo) or reordered names, without matching genuinely different people who
# happen to share a few letters (e.g. "Priya Singh" vs "Priya Shah" scores well
# below this).
FUZZY_MATCH_THRESHOLD = 0.86


def _match_fuzzy(en, p_norm, used):
    """Best whole-name similarity match above FUZZY_MATCH_THRESHOLD, or None.
    Catches typos and reordered names that share no exact token with _match_partial."""
    best_p, best_i, best_score = None, None, 0.0
    for i, (pn, p) in enumerate(p_norm):
        if i in used or not pn:
            continue
        score = difflib.SequenceMatcher(None, en, pn).ratio()
        if score > best_score:
            best_p, best_i, best_score = p, i, score
    if best_score >= FUZZY_MATCH_THRESHOLD:
        return best_p, best_i
    return None, None


def _match_email(en_email, participants_by_email, used):
    """Exact (case-insensitive) email match — used when the Teams display name
    doesn't resemble the enrolled name at all (nicknames, personal devices, etc.)
    but both sides recorded the same email/UPN."""
    if not en_email:
        return None, None
    match = participants_by_email.get(_norm(en_email))
    if match is None:
        return None, None
    p, i = match
    if i in used:
        return None, None
    return p, i


def _match_one_enrolled(en, en_parts, en_email, p_norm, participants_by_email, used):
    """Returns (match, used_i, uncertain) trying, in order: exact name -> partial
    name (token containment) -> email -> fuzzy name similarity -> first-name-only.
    Email is checked before fuzzy/first-name since a matching email is a far more
    reliable identity signal than a loose name heuristic."""
    match, used_i = _match_exact(en, p_norm, used)
    if match is not None:
        return match, used_i, False
    match, used_i = _match_partial(en_parts, p_norm, used)
    if match is not None:
        return match, used_i, False
    match, used_i = _match_email(en_email, participants_by_email, used)
    if match is not None:
        return match, used_i, False
    match, used_i = _match_fuzzy(en, p_norm, used)
    if match is not None:
        return match, used_i, True
    match, used_i = _match_first_name(en_parts, p_norm, used)
    if match is not None:
        return match, used_i, True
    return None, None, False


def _result_for_match(match, session_minutes, threshold_pct, max_minutes, uncertain):
    present = (match["minutes"] / session_minutes) >= (threshold_pct / 100.0) if session_minutes else False
    # Attentiveness = participant duration / highest duration among all participants,
    # rounded to whole-percent granularity (no decimal points in the report).
    attentiveness = round(match["minutes"] / max_minutes, 2) if max_minutes else 0.0
    return {
        "present": present,
        "attentiveness": attentiveness,
        "matched": True,
        "uncertain": uncertain,
        "minutes": match["minutes"],
    }


def match_participants(enrolled_names, participants, session_minutes, threshold_pct,
                       max_minutes=0.0, enrolled_emails=None):
    """Returns (results_by_enrolled, unmatched_participants).
    results_by_enrolled[name] = {present, attentiveness, matched, uncertain}
    enrolled_emails: optional {enrolled_name: email}, from an email column in the
    consolidated report, used as a matching fallback when name-matching fails."""
    enrolled_emails = enrolled_emails or {}
    results = {}
    used = set()
    p_norm = [(_norm(p["name"]), p) for p in participants]
    participants_by_email = {}
    for i, p in enumerate(participants):
        em = _norm(p.get("email", ""))
        if em:
            participants_by_email[em] = (p, i)

    for enrolled in enrolled_names:
        en = _norm(enrolled)
        en_parts = en.split()
        en_email = enrolled_emails.get(enrolled, "")
        match, used_i, uncertain = _match_one_enrolled(
            en, en_parts, en_email, p_norm, participants_by_email, used)
        if match is not None:
            used.add(used_i)
            results[enrolled] = _result_for_match(match, session_minutes, threshold_pct, max_minutes, uncertain)
        else:
            # Absent participant -> 0 (matches existing sheet's absent format)
            results[enrolled] = {"present": False, "attentiveness": 0.0,
                                 "matched": False, "uncertain": False, "minutes": 0}

    unmatched = [p["raw"] for i, (pn, p) in enumerate(p_norm) if i not in used]
    return results, unmatched


# ----------------------------- Sheet helpers -----------------------------
def _find_name_column(ws, max_scan_rows=3):
    for r in range(1, max_scan_rows + 1):
        for c in range(1, min(ws.max_column, 10) + 1):
            v = ws.cell(row=r, column=c).value
            if v and "name" in str(v).lower():
                return c
    return 2  # default column B


def _find_email_column(ws, max_scan_rows=3):
    """Column index of an Email/UPN header in this sheet, or None if it doesn't
    have one — trackers without an email column simply skip email-based matching."""
    for r in range(1, max_scan_rows + 1):
        for c in range(1, min(ws.max_column, 15) + 1):
            v = str(ws.cell(row=r, column=c).value or "").strip().lower()
            if v and ("email" in v or "upn" in v):
                return c
    return None


def _enrolled_rows(ws, name_col, start_row):
    rows = []
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=name_col).value
        if v and str(v).strip():
            rows.append((r, str(v).strip()))
    return rows


_SUBHEADER_TOKENS = {"attendance", "attentiveness", "overall"}


def _find_subheader_row(ws, max_col, default=2):
    """Trackers vary: some put Attendance/Attentiveness sub-headers on row 2 (row 3 is
    the first data row), others leave row 2 blank and use row 3 (data starts row 4).
    Detect the real row from existing columns instead of assuming — guessing wrong
    means new columns get written to the wrong row and never find a style template.
    Uses a majority vote across columns rather than the first match, so a single
    stray label (e.g. from an earlier mis-appended column) can't throw it off."""
    counts = {}
    for r in range(2, 5):
        n = sum(
            1 for c in range(1, max_col + 1)
            if str(ws.cell(row=r, column=c).value or "").strip().lower() in _SUBHEADER_TOKENS
        )
        if n:
            counts[r] = n
    return max(counts, key=counts.get) if counts else default


# ----------------------------- Alert highlighting -----------------------------
# A participant who's absent, or present but barely attentive, should stand out —
# but different trackers already carry their own idea of "the alert color" (some
# bake a static red fill into cells, others rely purely on Excel conditional-
# formatting rules scoped to old columns that never reaches newly appended ones).
# Prefer whatever this sheet is already using; only fall back to a default when
# the sheet has no baked color to find.
LOW_ATTENTIVENESS_THRESHOLD = 0.50
_DEFAULT_ALERT_FILL = PatternFill("solid", fgColor="FFC7CE")
_DEFAULT_ALERT_FONT_COLOR = "9C0006"


def _find_alert_fill(ws, col, rows):
    """Scan an existing Attendance/Attentiveness column for a row that already
    carries a baked (non-conditional-formatting) solid fill, so a tracker with
    its own established highlight color keeps that exact color going forward."""
    if not col:
        return None, None
    for r, _name in rows:
        cell = ws.cell(row=r, column=col)
        fill = cell.fill
        if fill and fill.patternType == "solid" and fill.fgColor:
            rgb = fill.fgColor.rgb
            if rgb and rgb not in (None, "00000000"):
                font_color = cell.font.color.rgb if cell.font and cell.font.color else None
                return copy(fill), font_color
    return None, None


def _clear_alert_style(cell):
    """Reset fill/font-color to a clean baseline before conditionally re-applying
    the alert style. copy_cell_style() pulls its baseline from the same row of the
    previous session's column — if that participant was flagged red on a prior day,
    the fill/font color would otherwise bleed into today's cell even when today's
    value doesn't warrant it. Non-color font attributes (bold, size, name) are kept."""
    cell.fill = PatternFill(fill_type=None)
    f = cell.font
    if f is not None:
        cell.font = Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                         vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
                         color=None)


def _header_is_merged(ws, col1, col2, row=1):
    """Whether an existing day-header spans col1..col2 as one merged cell — so a
    newly appended header only gets merged if that's this tracker's own convention."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col1 and rng.max_col >= col2:
            return True
    return False


# ----------------------------- Append logic: Consolidated Report -----------------------------
def _find_consolidated_template_cols(ws, sub_row, max_col):
    tmpl_att = tmpl_atten = None
    for c in range(max_col, 1, -1):
        val = ws.cell(row=sub_row, column=c).value
        if val and str(val).strip().lower() == "attentiveness":
            tmpl_atten = c
            tmpl_att = c - 1
            break
    if tmpl_att is None:
        for c in range(max_col, 1, -1):
            if str(ws.cell(row=sub_row, column=c).value or "").strip().lower() == "attendance":
                tmpl_att = c
                tmpl_atten = c
                break
    return tmpl_att, tmpl_atten


def _write_consolidated_header(ws, sub_row, new_att, new_atten, tmpl_att, tmpl_atten, header_label):
    # Row 1 header — merged across both new columns only if that's how this
    # tracker's existing day-headers are already set up (skip merging otherwise).
    if tmpl_att and tmpl_atten and new_att != new_atten and _header_is_merged(ws, tmpl_att, tmpl_atten):
        ws.merge_cells(start_row=1, start_column=new_att, end_row=1, end_column=new_atten)
    h = ws.cell(row=1, column=new_att, value=header_label)
    if tmpl_att:
        copy_cell_style(ws.cell(row=1, column=tmpl_att), h)
    # sub headers, on whichever row this tracker actually uses
    a2 = ws.cell(row=sub_row, column=new_att, value="Attendance")
    b2 = ws.cell(row=sub_row, column=new_atten, value="Attentiveness")
    if tmpl_att:
        copy_cell_style(ws.cell(row=sub_row, column=tmpl_att), a2)
    if tmpl_atten:
        copy_cell_style(ws.cell(row=sub_row, column=tmpl_atten), b2)


def _write_consolidated_row(ws, r, name, results, new_att, new_atten, tmpl_att, tmpl_atten,
                            alert_fill, alert_font_color):
    res = results.get(name, {})
    present = res.get("present")
    att = res.get("attentiveness", 0.0) or 0.0
    ac = ws.cell(row=r, column=new_att, value="Yes" if present else "No")
    bc = ws.cell(row=r, column=new_atten, value=att)
    if tmpl_att:
        copy_cell_style(ws.cell(row=r, column=tmpl_att), ac)
    if tmpl_atten:
        copy_cell_style(ws.cell(row=r, column=tmpl_atten), bc)
    _clear_alert_style(ac)
    _clear_alert_style(bc)
    bc.number_format = "0%"  # display computed attentiveness as a whole-number percentage

    # Flag absent, or present-but-disengaged (attentiveness <= threshold) rows —
    # highlight both cells so a "Yes" that was barely paying attention is as
    # visible as an outright "No".
    if not present or att <= LOW_ATTENTIVENESS_THRESHOLD:
        ac.fill = copy(alert_fill)
        bc.fill = copy(alert_fill)
        if alert_font_color:
            ac.font = Font(color=alert_font_color, bold=ac.font.bold if ac.font else False)
            bc.font = Font(color=alert_font_color, bold=bc.font.bold if bc.font else False)


def _append_consolidated(ws, header_label, results):
    name_col = _find_name_column(ws)
    max_col = ws.max_column
    sub_row = _find_subheader_row(ws, max_col)
    rows = _enrolled_rows(ws, name_col, sub_row + 1)

    tmpl_att, tmpl_atten = _find_consolidated_template_cols(ws, sub_row, max_col)

    # existing sessions are separated by a single blank gap column; match that layout
    gap_col = max_col + 1
    new_att = max_col + 2
    new_atten = max_col + 3
    if tmpl_att:
        copy_column_width(ws, tmpl_att - 1 if tmpl_att > 1 else tmpl_att, gap_col)
        copy_column_width(ws, tmpl_att, new_att)
    if tmpl_atten:
        copy_column_width(ws, tmpl_atten, new_atten)

    _write_consolidated_header(ws, sub_row, new_att, new_atten, tmpl_att, tmpl_atten, header_label)

    alert_fill, alert_font_color = _find_alert_fill(ws, tmpl_att, rows)
    if alert_fill is None:
        alert_fill = _DEFAULT_ALERT_FILL
        alert_font_color = _DEFAULT_ALERT_FONT_COLOR

    for r, name in rows:
        _write_consolidated_row(ws, r, name, results, new_att, new_atten, tmpl_att, tmpl_atten,
                                alert_fill, alert_font_color)


# ----------------------------- Append logic: Overall Attendance % -----------------------------
def _find_overall_template_col(ws, sub_row, max_col):
    for c in range(max_col, 1, -1):
        if str(ws.cell(row=sub_row, column=c).value or "").strip().lower() == "attendance":
            return c
    return None


def _write_overall_row(ws, r, name, results, new_col, tmpl_col, alert_fill, alert_font_color):
    res = results.get(name, {})
    present = res.get("present")
    cell = ws.cell(row=r, column=new_col, value="Yes" if present else "No")
    if tmpl_col:
        copy_cell_style(ws.cell(row=r, column=tmpl_col), cell)
    _clear_alert_style(cell)
    if not present:
        cell.fill = copy(alert_fill)
        if alert_font_color:
            cell.font = Font(color=alert_font_color, bold=cell.font.bold if cell.font else False)


def _append_overall(ws, header_label, results):
    name_col = _find_name_column(ws)
    max_col = ws.max_column
    sub_row = _find_subheader_row(ws, max_col)
    rows = _enrolled_rows(ws, name_col, sub_row + 1)
    new_col = max_col + 1

    tmpl_col = _find_overall_template_col(ws, sub_row, max_col)

    h = ws.cell(row=1, column=new_col, value=header_label)
    sub = ws.cell(row=sub_row, column=new_col, value="Attendance")
    if tmpl_col:
        copy_cell_style(ws.cell(row=1, column=tmpl_col), h)
        copy_cell_style(ws.cell(row=sub_row, column=tmpl_col), sub)
        copy_column_width(ws, tmpl_col, new_col)

    alert_fill, alert_font_color = _find_alert_fill(ws, tmpl_col, rows)
    if alert_fill is None:
        alert_fill = _DEFAULT_ALERT_FILL
        alert_font_color = _DEFAULT_ALERT_FONT_COLOR

    for r, name in rows:
        _write_overall_row(ws, r, name, results, new_col, tmpl_col, alert_fill, alert_font_color)


# ----------------------------- Append logic: Login -----------------------------
def _find_login_template_start(ws):
    # groups: 4 cols + 1 gap, starting col 1 -> starts at 1,6,11,...
    # find rightmost group by scanning row 3 for 'Name'
    starts = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=3, column=c).value
        if v and str(v).strip().lower() == "name":
            starts.append(c)
    return starts[-1] if starts else None


def _write_login_participant_row(ws, r, p, cols, tmpl_start):
    vals = [sanitize_excel_value(p["raw"]), p.get("first_join", ""), p.get("last_leave", ""), _fmt_minutes(p["minutes"])]
    for i, val in enumerate(vals):
        cell = ws.cell(row=r, column=cols[i], value=val)
        if tmpl_start:
            copy_cell_style(ws.cell(row=4, column=tmpl_start + i), cell)


def _append_login(ws, header_label, participants):
    last_start = _find_login_template_start(ws)
    if last_start:
        new_start = last_start + 5
        tmpl_start = last_start
    else:
        new_start = 1
        tmpl_start = None
    cols = [new_start, new_start + 1, new_start + 2, new_start + 3]

    # row1-2 merged header
    ws.merge_cells(start_row=1, start_column=cols[0], end_row=2, end_column=cols[3])
    h = ws.cell(row=1, column=cols[0], value=header_label)
    if tmpl_start:
        copy_cell_style(ws.cell(row=1, column=tmpl_start), h)
    # row3 subheaders
    labels = ["Name", "First Join", "Last Leave", "In-Meeting Duration"]
    for i, lab in enumerate(labels):
        cell = ws.cell(row=3, column=cols[i], value=lab)
        if tmpl_start:
            copy_cell_style(ws.cell(row=3, column=tmpl_start + i), cell)
            copy_column_width(ws, tmpl_start + i, cols[i])
    # data rows from row 4 — copy style from the template group's first data row (row 4)
    for di, p in enumerate(participants):
        r = 4 + di
        _write_login_participant_row(ws, r, p, cols, tmpl_start)


def _fmt_minutes(mins):
    h = int(mins // 60)
    m = int(mins % 60)
    s = int(round((mins - int(mins)) * 60))
    parts = []
    if h:
        parts.append(f"{h}h")
    if m:
        parts.append(f"{m}m")
    if s:
        parts.append(f"{s}s")
    return " ".join(parts) if parts else "0m"


def _process_one_session(wb, teams_bytes, teams_name, session_name, session_date_iso, threshold_pct,
                         scheduled_minutes=None):
    """Appends a single Teams export's session onto an already-open tracker workbook
    (mutated in place) and returns that day's info dict. Shared by the single-session
    and multi-day-batch entry points so a batch is just this called N times against
    the same workbook before a single save, instead of N separate load/save round trips.

    scheduled_minutes: the session's planned duration from the program schedule, if
    known. Attendance/attentiveness are measured against whichever is SMALLER of the
    Teams-derived duration and this scheduled one — so a meeting room left open past
    the planned end doesn't inflate the denominator for everyone (capped at what was
    scheduled), while a session that wrapped up early is still measured against what
    actually happened (the shorter, real duration), not padded out to the schedule."""
    parsed = parse_teams_export(teams_bytes, teams_name)
    participants = parsed["participants"]
    session_minutes = parsed["session_minutes"]
    max_minutes = parsed["max_minutes"]
    capped = False
    if scheduled_minutes:
        if session_minutes > scheduled_minutes:
            session_minutes = scheduled_minutes
            capped = True
        if max_minutes > scheduled_minutes:
            max_minutes = scheduled_minutes
            capped = True

    sheetnames = wb.sheetnames

    # locate sheets by best match
    def find_sheet(keys, default_idx):
        for sn in sheetnames:
            low = sn.lower()
            if any(k in low for k in keys):
                return wb[sn]
        return wb[sheetnames[default_idx]] if default_idx < len(sheetnames) else None

    ws_consol = find_sheet(["consolidated", "report"], 0)
    ws_overall = find_sheet(["overall", "attendance %"], 1)
    ws_login = find_sheet(["login"], 2)

    enrolled = []
    enrolled_emails = {}
    if ws_consol is not None:
        nc = _find_name_column(ws_consol)
        ec = _find_email_column(ws_consol)
        sub_row = _find_subheader_row(ws_consol, ws_consol.max_column)
        rows = _enrolled_rows(ws_consol, nc, sub_row + 1)
        enrolled = [n for _, n in rows]
        if ec:
            for r, n in rows:
                v = ws_consol.cell(row=r, column=ec).value
                if v and str(v).strip():
                    enrolled_emails[n] = str(v).strip()

    results, unmatched = match_participants(
        enrolled, participants, session_minutes, threshold_pct, max_minutes, enrolled_emails)

    # date label like (22-05-26)
    from datetime import datetime
    dt = datetime.strptime(session_date_iso, "%Y-%m-%d")
    date_label = dt.strftime("%d-%m-%y")
    header_label = sanitize_excel_value(f"{session_name} ({date_label})")

    if ws_consol is not None:
        _append_consolidated(ws_consol, header_label, results)
    if ws_overall is not None:
        _append_overall(ws_overall, header_label, results)
    if ws_login is not None:
        _append_login(ws_login, header_label, participants)

    matched_atts = [r["attentiveness"] for r in results.values() if r.get("matched")]
    present_count = sum(1 for r in results.values() if r.get("present"))
    # preserve enrolled-list order for the absent-names list
    absent_names = [name for name in enrolled if not results.get(name, {}).get("present")]
    avg_attendance_pct = round((present_count / len(enrolled)) * 100) if enrolled else 0
    avg_attentiveness_pct = round((sum(matched_atts) / len(matched_atts)) * 100) if matched_atts else 0

    return {
        "unmatched": unmatched,
        "uncertain": [n for n, r in results.items() if r.get("uncertain")],
        "total_participants": len(participants),
        "matched": sum(1 for r in results.values() if r.get("matched")),
        "present": present_count,
        "absent": len(absent_names),
        "absent_names": absent_names,
        "enrolled": len(enrolled),
        "session_minutes": round(session_minutes, 1),
        "capped_by_schedule": capped,
        "avg_attendance_pct": avg_attendance_pct,
        "avg_attentiveness_pct": avg_attentiveness_pct,
    }


def process_attendance(tracker_bytes, tracker_name, teams_bytes, teams_name,
                       session_name, session_date_iso, threshold_pct, scheduled_minutes=None):
    """Main entry for a single session. Returns (output_bytes, output_filename, info_dict).
    scheduled_minutes: the program's planned duration for this date, if known — see
    _process_one_session for how it's used to cap the attendance/attentiveness denominator."""
    wb = openpyxl.load_workbook(io.BytesIO(tracker_bytes))
    info = _process_one_session(wb, teams_bytes, teams_name, session_name, session_date_iso, threshold_pct,
                                scheduled_minutes)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    out_name = rename_with_new_date(tracker_name, session_date_iso)
    info["output_filename"] = out_name
    return out.getvalue(), out_name, info


def process_attendance_batch(tracker_bytes, tracker_name, sessions, threshold_pct,
                             feedback_bytes=None, sessions_by_date=None, scheduled_minutes_by_date=None):
    """Consolidates several days' Teams exports into one tracker in a single pass.
    sessions: list of {bytes, filename, session_name, session_date_iso}. Processed in
    date order (regardless of upload order) so the tracker's appended columns read
    left-to-right chronologically. If feedback_bytes is given, a raw feedback-form
    export (one row per submission, no date/module/mentor columns) is also
    consolidated into the tracker's Feedback sheet, using sessions_by_date
    ({iso_date: {module_name, faculty}}, from the program's schedule) to fill in
    the date/module/mentor for each submission and skipping dates already captured.
    scheduled_minutes_by_date: optional {iso_date: minutes} — each day's planned
    duration from the program schedule, used to cap that day's attendance/
    attentiveness denominator (see _process_one_session).
    Returns (output_bytes, output_filename, summary_dict)."""
    if not sessions:
        raise ValueError("No sessions to process")

    ordered = sorted(sessions, key=lambda s: s["session_date_iso"])
    wb = openpyxl.load_workbook(io.BytesIO(tracker_bytes))
    scheduled_minutes_by_date = scheduled_minutes_by_date or {}

    days = []
    for s in ordered:
        info = _process_one_session(
            wb, s["bytes"], s["filename"], s["session_name"], s["session_date_iso"], threshold_pct,
            scheduled_minutes_by_date.get(s["session_date_iso"]))
        info["session_name"] = s["session_name"]
        info["session_date"] = s["session_date_iso"]
        info["source_filename"] = s["filename"]
        days.append(info)

    summary = {
        "days": days,
        "sessions_processed": len(days),
    }

    if feedback_bytes is not None:
        feedback_rows = parse_feedback_export(feedback_bytes)
        summary["feedback"] = append_feedback_sheet(wb, feedback_rows, sessions_by_date or {})

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    latest_date = ordered[-1]["session_date_iso"]
    out_name = rename_with_new_date(tracker_name, latest_date)
    summary["output_filename"] = out_name
    return out.getvalue(), out_name, summary
