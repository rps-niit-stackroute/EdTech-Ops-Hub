"""Unit tests for feedback_processor.py — raw feedback-form export consolidation."""
import io
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from feedback_processor import (
    _norm, _find_col, _header_map, _as_datetime, parse_feedback_export,
    _find_feedback_sheet, _feedback_columns, _scan_existing_feedback_rows,
    _row_values, _write_feedback_row, _process_feedback_row, append_feedback_sheet,
)


class TestNorm:
    def test_collapses_whitespace_and_lowercases(self):
        assert _norm("  Participant   Name  ") == "participant name"

    def test_none_returns_empty(self):
        assert _norm(None) == ""


class TestFindCol:
    def test_all_substrings_must_match(self):
        header_map = {"participant name": 1, "other feedback": 2}
        assert _find_col(header_map, "participant", "name") == 1
        assert _find_col(header_map, "specific", "feedback") is None

    def test_no_match_returns_none(self):
        assert _find_col({"a": 1}, "b") is None


def _wb_with_headers(headers, data_rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in (data_rows or []):
        ws.append(row)
    return wb, ws


class TestHeaderMap:
    def test_builds_normalized_map(self):
        wb, ws = _wb_with_headers(["Participant Name", "Rating"])
        m = _header_map(ws)
        assert m["participant name"] == 1
        assert m["rating"] == 2

    def test_skips_blank_cells(self):
        wb, ws = _wb_with_headers(["Name", None, "Rating"])
        m = _header_map(ws)
        assert 2 not in m.values()


class TestAsDatetime:
    def test_datetime_passthrough(self):
        dt = datetime(2026, 5, 22)
        assert _as_datetime(dt) is dt

    def test_non_datetime_returns_none(self):
        assert _as_datetime("2026-05-22") is None
        assert _as_datetime(None) is None


class TestParseFeedbackExport:
    def test_parses_rows_with_participant_name(self):
        wb, ws = _wb_with_headers(
            ["Participant Name", "Completion time", "Takeaways", "Rating",
             "Specific feedback", "Any other feedback"],
            [["John Doe", datetime(2026, 5, 22, 10, 0), "Learned a lot", 5, "Great pace", "None"]],
        )
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_feedback_export(buf.getvalue())
        assert len(rows) == 1
        r = rows[0]
        assert r["participant_name"] == "John Doe"
        assert r["dt"] == datetime(2026, 5, 22, 10, 0)
        assert r["rating"] == 5

    def test_blank_participant_skipped(self):
        wb, ws = _wb_with_headers(
            ["Participant Name", "Completion time"],
            [["", datetime(2026, 5, 22)], ["  ", datetime(2026, 5, 22)]],
        )
        buf = io.BytesIO()
        wb.save(buf)
        assert parse_feedback_export(buf.getvalue()) == []

    def test_falls_back_to_start_time_when_no_completion_time(self):
        wb, ws = _wb_with_headers(
            ["Participant Name", "Start time"],
            [["John", datetime(2026, 5, 22)]],
        )
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_feedback_export(buf.getvalue())
        assert rows[0]["dt"] == datetime(2026, 5, 22)

    def test_rating_column_fallback_chain(self):
        # No "scale"/"rating" column, but an "effective" column should be used.
        wb, ws = _wb_with_headers(
            ["Participant Name", "How effective was this session"],
            [["John", 4]],
        )
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_feedback_export(buf.getvalue())
        assert rows[0]["rating"] == 4


class TestFindFeedbackSheet:
    def test_finds_sheet_by_name_substring(self):
        wb = openpyxl.Workbook()
        wb.create_sheet("Feedback Sheet")
        assert _find_feedback_sheet(wb) is not None
        assert _find_feedback_sheet(wb).title == "Feedback Sheet"

    def test_returns_none_when_absent(self):
        wb = openpyxl.Workbook()
        assert _find_feedback_sheet(wb) is None


class TestFeedbackColumns:
    def test_defaults_sno_to_1(self):
        cols = _feedback_columns({})
        assert cols["sno"] == 1

    def test_maps_all_expected_keys(self):
        header_map = {"sno": 0, "date": 1, "module": 2, "faculty": 3,
                       "participant name": 4, "takeaway": 5, "rating": 6,
                       "specific feedback": 7, "any other feedback": 8}
        cols = _feedback_columns(header_map)
        assert cols["date"] == 1
        assert cols["module"] == 2
        assert cols["faculty"] == 3
        assert cols["participant"] == 4


class TestScanExistingFeedbackRows:
    def test_empty_sheet(self):
        wb, ws = _wb_with_headers(["SNo", "Date"])
        cols = {"sno": 1, "date": 2}
        dates, max_sno, last_row = _scan_existing_feedback_rows(ws, cols)
        assert dates == set()
        assert max_sno == 0
        assert last_row == 1

    def test_tracks_max_sno_and_dates(self):
        wb, ws = _wb_with_headers(
            ["SNo", "Date"],
            [[1, datetime(2026, 5, 20)], [2, datetime(2026, 5, 21)]],
        )
        cols = {"sno": 1, "date": 2}
        dates, max_sno, last_row = _scan_existing_feedback_rows(ws, cols)
        assert dates == {"2026-05-20", "2026-05-21"}
        assert max_sno == 2
        assert last_row == 3

    def test_blank_rows_ignored(self):
        wb, ws = _wb_with_headers(["SNo", "Date"], [[None, None]])
        cols = {"sno": 1, "date": 2}
        dates, max_sno, last_row = _scan_existing_feedback_rows(ws, cols)
        assert last_row == 1


class TestRowValues:
    def test_maps_row_and_session_into_column_positions(self):
        row = {"dt": datetime(2026, 5, 22), "participant_name": "John",
               "takeaways": "Good", "rating": 5, "specific_feedback": "x", "other_feedback": "y"}
        session = {"module_name": "Intro", "faculty": "Jane"}
        cols = {"sno": 1, "date": 2, "module": 3, "faculty": 4, "participant": 5,
                "takeaways": 6, "rating": 7, "specific": 8, "other": 9}
        values = _row_values(row, session, cols, 3)
        assert values[1] == 3
        assert values[3] == "Intro"
        assert values[4] == "Jane"
        assert values[5] == "John"


def _feedback_wb():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback"
    ws.append(["SNo", "Date", "Module", "Faculty", "Participant Name",
               "Takeaways", "Rating", "Specific Feedback", "Any Other Feedback"])
    return wb, ws


class TestProcessFeedbackRow:
    def test_skips_already_existing_date(self):
        wb, ws = _feedback_wb()
        row = {"dt": datetime(2026, 5, 20), "participant_name": "John"}
        status, iso = _process_feedback_row(
            ws, row, {"sno": 1, "date": 2}, {}, {"2026-05-20"}, None, 1, 2)
        assert status == "skipped_existing"
        assert iso == "2026-05-20"

    def test_unmatched_when_no_session_for_date(self):
        wb, ws = _feedback_wb()
        row = {"dt": datetime(2026, 5, 20), "participant_name": "John"}
        status, iso = _process_feedback_row(
            ws, row, {"sno": 1, "date": 2}, {}, set(), None, 1, 2)
        assert status == "unmatched"

    def test_added_when_session_found(self):
        wb, ws = _feedback_wb()
        row = {"dt": datetime(2026, 5, 20), "participant_name": "John", "takeaways": "",
               "rating": 5, "specific_feedback": "", "other_feedback": ""}
        cols = {"sno": 1, "date": 2, "module": 3, "faculty": 4, "participant": 5,
                "takeaways": 6, "rating": 7, "specific": 8, "other": 9}
        sessions_by_date = {"2026-05-20": {"module_name": "Intro", "faculty": "Jane"}}
        status, iso = _process_feedback_row(ws, row, cols, sessions_by_date, set(), None, 1, 2)
        assert status == "added"
        assert ws.cell(row=2, column=3).value == "Intro"
        assert ws.cell(row=2, column=5).value == "John"


class TestAppendFeedbackSheet:
    def test_no_feedback_sheet_found(self):
        wb = openpyxl.Workbook()
        result = append_feedback_sheet(wb, [], {})
        assert result["sheet_found"] is False
        assert result["added"] == 0

    def test_missing_required_columns(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Feedback"
        ws.append(["Notes"])  # no Date or Participant Name column
        result = append_feedback_sheet(wb, [], {})
        assert result["sheet_found"] is True
        assert result["columns_missing"] is True

    def test_full_consolidation_flow(self):
        wb, ws = _feedback_wb()
        feedback_rows = [
            {"dt": datetime(2026, 5, 20), "participant_name": "John", "takeaways": "Good",
             "rating": 5, "specific_feedback": "", "other_feedback": ""},
            {"dt": datetime(2026, 5, 21), "participant_name": "Jane", "takeaways": "",
             "rating": 4, "specific_feedback": "", "other_feedback": ""},
            {"dt": None, "participant_name": "NoDate", "takeaways": "", "rating": None,
             "specific_feedback": "", "other_feedback": ""},
        ]
        sessions_by_date = {
            "2026-05-20": {"module_name": "Intro", "faculty": "Jane F"},
        }
        result = append_feedback_sheet(wb, feedback_rows, sessions_by_date)
        assert result["sheet_found"] is True
        assert result["added"] == 1
        assert result["unmatched_rows"] == 1
        assert result["unmatched_dates"] == ["2026-05-21"]
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=5).value == "John"

    def test_skips_dates_already_captured(self):
        wb, ws = _feedback_wb()
        ws.append([1, datetime(2026, 5, 20), "Intro", "Jane", "Existing", "", 5, "", ""])
        feedback_rows = [
            {"dt": datetime(2026, 5, 20), "participant_name": "John", "takeaways": "",
             "rating": 5, "specific_feedback": "", "other_feedback": ""},
        ]
        result = append_feedback_sheet(wb, feedback_rows, {"2026-05-20": {"module_name": "Intro", "faculty": "Jane"}})
        assert result["added"] == 0
        assert result["skipped_existing_dates"] == 1

    def test_new_rows_continue_sno_sequence(self):
        wb, ws = _feedback_wb()
        ws.append([1, datetime(2026, 5, 19), "Intro", "Jane", "Prev", "", 5, "", ""])
        feedback_rows = [
            {"dt": datetime(2026, 5, 20), "participant_name": "John", "takeaways": "",
             "rating": 5, "specific_feedback": "", "other_feedback": ""},
        ]
        append_feedback_sheet(wb, feedback_rows, {"2026-05-20": {"module_name": "Intro", "faculty": "Jane"}})
        assert ws.cell(row=3, column=1).value == 2
