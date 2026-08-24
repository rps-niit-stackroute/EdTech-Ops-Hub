"""Unit tests for excel_utils.py — pure date/time/string helpers, no I/O or DB."""
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from excel_utils import (
    ordinal, parse_time_to_24h, time_to_minutes, parse_date_flexible,
    parse_duration_to_minutes, sanitize_excel_value, copy_cell_style,
    normalize_teams_name, rename_with_new_date, copy_column_width,
)


class TestOrdinal:
    @pytest.mark.parametrize("n,expected", [
        (1, "st"), (2, "nd"), (3, "rd"), (4, "th"), (11, "th"), (12, "th"),
        (13, "th"), (20, "th"), (21, "st"), (22, "nd"), (23, "rd"), (24, "th"),
        (100, "th"), (101, "st"), (111, "th"),
    ])
    def test_ordinal_suffix(self, n, expected):
        assert ordinal(n) == expected


class TestParseTimeTo24h:
    def test_none_input(self):
        assert parse_time_to_24h(None) is None

    def test_empty_string(self):
        assert parse_time_to_24h("") is None
        assert parse_time_to_24h("   ") is None

    def test_datetime_object(self):
        dt = datetime(2026, 5, 22, 14, 30)
        assert parse_time_to_24h(dt) == "14:30"

    def test_plain_24h(self):
        assert parse_time_to_24h("10:00") == "10:00"

    def test_am_pm_with_colon(self):
        assert parse_time_to_24h("10:00 AM") == "10:00"
        assert parse_time_to_24h("12:00 PM") == "12:00"
        assert parse_time_to_24h("12:00 AM") == "00:00"
        assert parse_time_to_24h("6 PM") == "18:00"
        assert parse_time_to_24h("6PM") == "18:00"

    def test_ist_suffix_stripped(self):
        assert parse_time_to_24h("10:00 AM (IST)") == "10:00"

    def test_no_minutes_defaults_zero(self):
        assert parse_time_to_24h("9 AM") == "09:00"

    def test_invalid_hour_returns_none(self):
        assert parse_time_to_24h("99:00") is None

    def test_unparseable_returns_none(self):
        assert parse_time_to_24h("not a time") is None


class TestTimeToMinutes:
    def test_none_and_empty(self):
        assert time_to_minutes(None) is None
        assert time_to_minutes("") is None

    def test_normal(self):
        assert time_to_minutes("01:30") == 90
        assert time_to_minutes("00:00") == 0
        assert time_to_minutes("23:59") == 23 * 60 + 59

    def test_malformed_returns_none(self):
        assert time_to_minutes("not-a-time") is None
        assert time_to_minutes("12") is None


class TestParseDateFlexible:
    def test_none(self):
        assert parse_date_flexible(None) is None

    def test_empty_string(self):
        assert parse_date_flexible("") is None
        assert parse_date_flexible("   ") is None

    def test_datetime_and_date_objects(self):
        assert parse_date_flexible(datetime(2026, 5, 22)) == "2026-05-22"
        assert parse_date_flexible(date(2026, 5, 22)) == "2026-05-22"

    @pytest.mark.parametrize("s,expected", [
        ("2026-05-22", "2026-05-22"),
        ("22-May-2026", "2026-05-22"),
        ("22-May-26", "2026-05-22"),
        ("22/05/2026", "2026-05-22"),
        ("05/22/2026", "2026-05-22"),
        ("22-05-2026", "2026-05-22"),
        ("22 May 2026", "2026-05-22"),
        ("2026/05/22", "2026-05-22"),
    ])
    def test_known_formats(self, s, expected):
        assert parse_date_flexible(s) == expected

    def test_ordinal_fallback(self):
        assert parse_date_flexible("20th May 2026") == "2026-05-20"

    def test_ordinal_fallback_no_year_uses_current_year(self):
        result = parse_date_flexible("20th May")
        assert result == f"{datetime.now().year}-05-20"

    def test_unparseable_returns_none(self):
        assert parse_date_flexible("gibberish text") is None

    def test_ordinal_fallback_invalid_month_returns_none(self):
        assert parse_date_flexible("20th Blah 2026") is None

    def test_ordinal_fallback_invalid_day_returns_none(self):
        # Feb 30 doesn't exist -> date() raises -> caught, returns None
        assert parse_date_flexible("30th Feb 2026") is None


class TestParseDurationToMinutes:
    def test_none_and_empty(self):
        assert parse_duration_to_minutes(None) == 0.0
        assert parse_duration_to_minutes("") == 0.0
        assert parse_duration_to_minutes("   ") == 0.0

    def test_hh_mm_ss_format(self):
        assert parse_duration_to_minutes("01:07:35") == pytest.approx(67.5833, abs=0.001)
        assert parse_duration_to_minutes("0:05:00") == 5.0

    def test_token_style(self):
        assert parse_duration_to_minutes("1.5h") == 90.0
        assert parse_duration_to_minutes("1h 7m 35s") == pytest.approx(67.5833, abs=0.001)
        assert parse_duration_to_minutes("55m 13s") == pytest.approx(55.2166, abs=0.001)
        assert parse_duration_to_minutes("3h") == 180.0
        assert parse_duration_to_minutes("2h 59m 52s") == pytest.approx(179.8666, abs=0.001)
        assert parse_duration_to_minutes("2.5h 30m") == 180.0

    def test_minutes_suffix_not_confused_with_seconds(self):
        # "m(?!s)" must not match the "m" in "ms"
        assert parse_duration_to_minutes("5s") == pytest.approx(5 / 60.0, abs=0.001)

    def test_plain_number_is_minutes(self):
        assert parse_duration_to_minutes("45") == 45.0
        assert parse_duration_to_minutes(45) == 45.0

    def test_unparseable_returns_zero(self):
        assert parse_duration_to_minutes("gibberish") == 0.0


class TestSanitizeExcelValue:
    @pytest.mark.parametrize("trigger", ["=", "+", "-", "@", "\t", "\r"])
    def test_formula_trigger_chars_get_quoted(self, trigger):
        v = f"{trigger}SUM(A1:A2)"
        assert sanitize_excel_value(v) == "'" + v

    def test_normal_string_untouched(self):
        assert sanitize_excel_value("John Doe") == "John Doe"

    def test_non_string_passthrough(self):
        assert sanitize_excel_value(42) == 42
        assert sanitize_excel_value(None) is None


class TestCopyCellStyle:
    def test_copies_style_attributes(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        src = ws.cell(row=1, column=1, value="x")
        src.font = Font(bold=True)
        dst = ws.cell(row=1, column=2, value="y")
        copy_cell_style(src, dst)
        assert dst.font.bold is True

    def test_swallows_exceptions(self):
        # Passing objects without the expected attributes must not raise.
        class Fake:
            pass
        copy_cell_style(Fake(), Fake())


class TestNormalizeTeamsName:
    def test_empty_and_none(self):
        assert normalize_teams_name("") == ""
        assert normalize_teams_name(None) == ""

    def test_last_comma_first_with_company(self):
        assert normalize_teams_name("Saleem, Abdul Baseer (Cognizant)") == "Abdul Baseer Saleem"

    def test_no_comma_passthrough(self):
        assert normalize_teams_name("John Doe") == "John Doe"

    def test_company_stripped_without_comma(self):
        assert normalize_teams_name("John Doe (Acme Corp)") == "John Doe"

    def test_collapses_extra_whitespace(self):
        assert normalize_teams_name("John    Doe") == "John Doe"


class TestRenameWithNewDate:
    def test_invalid_new_date_returns_original(self):
        assert rename_with_new_date("tracker.xlsx", "not-a-date") == "tracker.xlsx"

    def test_month_name_with_ordinal(self):
        assert rename_with_new_date("NIIT_CNA_20thMay.xlsx", "2026-05-22") == "NIIT_CNA_22ndMay.xlsx"

    def test_month_name_uppercase_ordinal_not_recognized(self):
        # The ordinal-suffix group only matches lowercase "st/nd/rd/th", so an
        # uppercase "TH" gets absorbed into the month-name token instead ("THMAY"),
        # which isn't a real month — the match is a no-op substitution, so
        # rename_with_new_date returns the filename unchanged rather than falling
        # through to append a new date.
        assert rename_with_new_date("Report_20THMAY.xlsx", "2026-06-05") == "Report_20THMAY.xlsx"

    def test_numeric_day_first(self):
        assert rename_with_new_date("tracker_22-05-26.xlsx", "2026-06-15") == "tracker_15-06-26.xlsx"

    def test_iso_date_segment(self):
        assert rename_with_new_date("tracker_2026-05-22.xlsx", "2026-06-15") == "tracker_2026-06-15.xlsx"

    def test_numeric_month_first_ambiguous_resolved_by_values(self):
        # day=13 > 12 so day-first is unambiguous
        assert rename_with_new_date("f_13-05-2026.xlsx", "2026-07-01") == "f_01-07-2026.xlsx"

    def test_no_recognizable_date_appends(self):
        result = rename_with_new_date("plain_file.xlsx", "2026-05-22")
        assert result == "plain_file_22-May-2026.xlsx"

    def test_no_extension_defaults_to_xlsx(self):
        result = rename_with_new_date("plain_file", "2026-05-22")
        assert result.endswith(".xlsx")

    def test_numeric_pattern_not_a_real_date_left_untouched(self):
        # 32 can't be a day or month in either order -> _numeric_day_first returns
        # None -> the replacement leaves the matched token as-is. That still counts
        # as a substitution (n > 0), so rename_with_new_date returns the filename
        # unchanged rather than falling through to the append-a-new-date branch.
        result = rename_with_new_date("code_32-13-2026.xlsx", "2026-01-01")
        assert result == "code_32-13-2026.xlsx"


class TestCopyColumnWidth:
    def test_copies_width_when_present(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions["A"].width = 25
        copy_column_width(ws, 1, 2)
        assert ws.column_dimensions["B"].width == 25

    def test_noop_when_source_has_no_width(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        copy_column_width(ws, 5, 6)
        # .get() (unlike subscript access) doesn't auto-vivify a ColumnDimension,
        # so this only passes if the function truly left column F untouched.
        assert ws.column_dimensions.get("F") is None
