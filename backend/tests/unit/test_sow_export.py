"""Unit tests for sow_export.py — SOW/Provision Excel generation."""
import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from sow_export import (
    _normalize_cell_value, _format_value, _load_previous_cells, _write_header,
    _write_row_with_diff, build_sow_excel, build_provision_excel, COLUMNS,
    PROVISION_COLUMNS,
)


class TestNormalizeCellValue:
    def test_none_becomes_empty_string(self):
        assert _normalize_cell_value(None) == ""

    def test_integer_float_loses_decimal(self):
        assert _normalize_cell_value(5.0) == "5"

    def test_non_integer_float_rounds_and_trims(self):
        assert _normalize_cell_value(5.10) == "5.1"

    def test_string_is_stripped(self):
        assert _normalize_cell_value("  hi  ") == "hi"


class TestFormatValue:
    def test_none_and_empty_string(self):
        assert _format_value(None) == ""
        assert _format_value("") == ""

    def test_integer_float(self):
        assert _format_value(3.0) == "3"

    def test_non_integer_float(self):
        assert _format_value(3.25) == "3.25"

    def test_other_values_stringified(self):
        assert _format_value(7) == "7"


class TestLoadPreviousCells:
    def test_none_bytes_returns_empty(self):
        assert _load_previous_cells(None) == {}

    def test_invalid_bytes_returns_empty(self):
        assert _load_previous_cells(b"not a real xlsx") == {}

    def test_loads_cell_values_by_position(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Header")
        ws.cell(row=2, column=1, value=42)
        buf = io.BytesIO()
        wb.save(buf)
        cells = _load_previous_cells(buf.getvalue())
        assert cells[(1, 1)] == "Header"
        assert cells[(2, 1)] == 42


class TestWriteHeader:
    def test_writes_all_column_headers_with_style(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        _write_header(ws, COLUMNS)
        assert ws.cell(row=1, column=1).value == "Month"
        assert ws.cell(row=1, column=1).font.bold is True
        assert ws.cell(row=1, column=len(COLUMNS)).value == COLUMNS[-1]


class TestWriteRowWithDiff:
    def test_no_previous_bytes_writes_plain_values(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        _write_row_with_diff(ws, 2, ["May", "John", 5.0], {}, None)
        assert ws.cell(row=2, column=1).value == "May"
        assert ws.cell(row=2, column=3).value == 5.0

    def test_unchanged_value_not_flagged(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        previous = {(2, 1): "May"}
        _write_row_with_diff(ws, 2, ["May"], previous, b"fake-bytes")
        assert ws.cell(row=2, column=1).value == "May"

    def test_changed_value_flagged_with_old_new(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        previous = {(2, 3): 3.0}
        _write_row_with_diff(ws, 2, ["May", "John", 5.0], previous, b"fake-bytes")
        cell = ws.cell(row=2, column=3)
        assert cell.value == "Old: 3 → New: 5"
        assert cell.fill.fgColor.rgb == "00FFFF00"


def _grouped_rows():
    return [{
        "mentor": "John Doe",
        "rows": [{
            "month": "May", "mentor": "John Doe", "total_hours": 5.0,
            "program_name": "Intro Program", "start_date": "1 May 2026",
            "end_date": "5 May 2026", "client": "Acme", "project_code": "P-001",
            "project_manager": "Santosh", "dates": "1 May, 5 May",
            "sessions_conducted": 2,
        }],
    }]


class TestBuildSowExcel:
    def test_builds_workbook_with_grand_total(self):
        xlsx_bytes = build_sow_excel(_grouped_rows(), "May 2026")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.title.startswith("SOW - May 2026")
        assert ws.cell(row=1, column=1).value == "Month"
        assert ws.cell(row=2, column=2).value == "John Doe"
        assert ws.cell(row=3, column=2).value == "GRAND TOTAL"
        assert ws.cell(row=3, column=3).value == 5.0
        assert ws.cell(row=3, column=11).value == 2

    def test_sheet_title_truncated_to_31_chars(self):
        long_label = "A" * 40
        xlsx_bytes = build_sow_excel([], long_label)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.active.title) <= 31

    def test_with_previous_sow_flags_changes(self):
        prev_wb = openpyxl.Workbook()
        prev_ws = prev_wb.active
        _write_header(prev_ws, COLUMNS)
        prev_ws.cell(row=2, column=3, value=3.0)  # old total_hours for the same row
        prev_ws.cell(row=3, column=3, value=3.0)  # old grand total hours
        prev_ws.cell(row=3, column=11, value=1)   # old grand total sessions
        buf = io.BytesIO()
        prev_wb.save(buf)

        xlsx_bytes = build_sow_excel(_grouped_rows(), "May 2026", previous_sow_bytes=buf.getvalue())
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert "Old: 3" in str(ws.cell(row=2, column=3).value)
        assert "New: 5" in str(ws.cell(row=2, column=3).value)


class TestBuildProvisionExcel:
    def test_builds_workbook_with_rows_and_charges(self):
        rows = [{
            "month": "May", "mentor": "Ashutosh", "total_hours": 4.0, "cost_per_hour": 1000,
            "total_cost": 4000, "program_name": "Prog", "start_date": "1 May 2026",
            "end_date": "2 May 2026", "client": "Acme", "project_code": "P-1",
            "project_manager": "Santosh", "dates": "1 May, 2 May", "sessions_conducted": 2,
        }]
        charges = [{
            "month_label": "May 2026", "trainer": "Vendor X", "description": "Retainer",
            "total_cost": 5000,
        }]
        xlsx_bytes = build_provision_excel(rows, charges, "May 2026")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.title.startswith("Provision - May 2026")
        assert ws.cell(row=1, column=1).value == "Month"
        assert ws.cell(row=2, column=2).value == "Ashutosh"
        assert ws.cell(row=3, column=2).value == "Vendor X"
        assert ws.cell(row=4, column=2).value == "GRAND TOTAL"
        assert ws.cell(row=4, column=3).value == 4.0
        assert ws.cell(row=4, column=5).value == 9000.0

    def test_no_rows_or_charges_still_produces_grand_total_row(self):
        xlsx_bytes = build_provision_excel([], [], "May 2026")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "GRAND TOTAL"
        assert ws.cell(row=2, column=3).value == 0.0
