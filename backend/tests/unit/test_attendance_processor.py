"""Unit tests for attendance_processor.py — Teams export parsing, name matching,
and tracker-append logic (Consolidated Report / Overall Attendance % / Login sheets)."""
import io
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from attendance_processor import (
    _decode_bytes, _rows_from_csv, _find_session_minutes, _find_participants_header,
    _parse_participant_rows, parse_teams_export, _detect_session_date,
    _parse_teams_datetime, _norm, _match_exact, _match_partial, _match_first_name,
    _match_fuzzy, _match_email, _find_email_column,
    match_participants, _find_name_column, _enrolled_rows, _find_subheader_row,
    _find_alert_fill, _header_is_merged, _fmt_minutes, process_attendance,
    process_attendance_batch,
)


def _teams_csv(rows_text, encoding="utf-16"):
    return rows_text.encode(encoding)


BASIC_EXPORT = (
    "Summary\n"
    "Meeting Start Time,5/22/26, 10:00:00 AM\n"
    "Meeting Duration,1h 30m\n"
    "\n"
    "3. In-Meeting Activities\n"
    "Name,First Join,Last Leave,In-Meeting Duration,Attentiveness\n"
    "\"Saleem, Abdul Baseer (Cognizant)\",5/22/26 10:00 AM,5/22/26 11:30 AM,1h 30m,\n"
    "John Doe,5/22/26 10:05 AM,5/22/26 11:00 AM,55m,\n"
    "Jane Roe,5/22/26 10:00 AM,5/22/26 10:20 AM,20m,\n"
)


class TestDecodeBytes:
    def test_utf16(self):
        assert _decode_bytes("hello".encode("utf-16")) == "hello"

    def test_utf8_sig(self):
        assert _decode_bytes("hello".encode("utf-8-sig")) is not None


class TestFindSessionMinutes:
    def test_finds_labeled_duration(self):
        rows = _rows_from_csv(_teams_csv("Meeting Duration,1h 30m\n"))
        assert _find_session_minutes(rows) == 90.0

    def test_no_label_returns_zero(self):
        rows = _rows_from_csv(_teams_csv("Name,Duration\nJohn,1h\n"))
        assert _find_session_minutes(rows) == 0.0

    def test_label_with_no_following_value_keeps_scanning(self):
        rows = _rows_from_csv(_teams_csv("Meeting Duration,\nMeeting Duration,45m\n"))
        assert _find_session_minutes(rows) == 45.0


class TestFindParticipantsHeader:
    def test_finds_header_with_name_and_duration(self):
        rows = _rows_from_csv(_teams_csv(
            "Name,First Join,Last Leave,In-Meeting Duration,Attentiveness\nJohn,1,2,3,4\n"))
        idx, cols = _find_participants_header(rows)
        assert idx == 0
        assert cols["name"] == 0
        assert cols["first_join"] == 1
        assert cols["last_leave"] == 2
        assert cols["duration"] == 3
        assert cols["attentiveness"] == 4

    def test_not_found_returns_none(self):
        rows = _rows_from_csv(_teams_csv("a,b\nc,d\n"))
        idx, cols = _find_participants_header(rows)
        assert idx is None
        assert cols == {}


class TestParseParticipantRows:
    def test_no_header_returns_empty(self):
        assert _parse_participant_rows([["a"]], None, {}) == []

    def test_missing_required_columns_returns_empty(self):
        assert _parse_participant_rows([["a"]], 0, {"name": 0}) == []

    def test_stops_at_next_section_marker(self):
        rows = [
            ["Name", "Duration"],
            ["John", "1h"],
            ["3. In-Meeting Activities"],
            ["Jane", "1h"],
        ]
        cols = {"name": 0, "duration": 1}
        participants = _parse_participant_rows(rows, 0, cols)
        assert len(participants) == 1
        assert participants[0]["raw"] == "John"

    def test_skips_blank_and_repeated_header_rows(self):
        rows = [
            ["Name", "Duration"],
            [],
            ["Name", "Duration"],
            ["John", "1h"],
        ]
        cols = {"name": 0, "duration": 1}
        participants = _parse_participant_rows(rows, 0, cols)
        assert len(participants) == 1


class TestParseTeamsExport:
    def test_full_export_csv(self):
        parsed = parse_teams_export(_teams_csv(BASIC_EXPORT), "export.csv")
        assert parsed["session_minutes"] == 90.0
        assert parsed["max_minutes"] == 90.0
        assert len(parsed["participants"]) == 3
        assert parsed["participants"][0]["name"] == "Abdul Baseer Saleem"
        assert parsed["session_date"] == "2026-05-22"

    def test_xlsx_input(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Meeting Duration", "1h"])
        ws.append(["Name", "First Join", "Last Leave", "In-Meeting Duration", "Attentiveness"])
        ws.append(["John Doe", "5/22/26 10:00 AM", "5/22/26 11:00 AM", "1h", ""])
        buf = io.BytesIO()
        wb.save(buf)
        parsed = parse_teams_export(buf.getvalue(), "export.xlsx")
        assert parsed["session_minutes"] == 60.0
        assert len(parsed["participants"]) == 1

    def test_session_minutes_falls_back_to_max_participant_minutes(self):
        text = (
            "3. In-Meeting Activities\n"
            "Name,First Join,Last Leave,In-Meeting Duration,Attentiveness\n"
            "John,1,2,45m,\n"
        )
        parsed = parse_teams_export(_teams_csv(text), "export.csv")
        assert parsed["session_minutes"] == 45.0


class TestDetectSessionDate:
    def test_from_meeting_start(self):
        rows = _rows_from_csv(_teams_csv("Meeting Start Time,5/22/26, 10:00 AM\n"))
        assert _detect_session_date(rows, []) == "2026-05-22"

    def test_falls_back_to_participant_first_join(self):
        # _parse_teams_datetime only splits on a comma before parsing, so the
        # date portion must be comma-separated from any trailing time text.
        rows = _rows_from_csv(_teams_csv("Name,Duration\nJohn,1h\n"))
        participants = [{"first_join": "5/23/26, 9:00 AM"}]
        assert _detect_session_date(rows, participants) == "2026-05-23"

    def test_no_candidates_returns_none(self):
        rows = _rows_from_csv(_teams_csv("a,b\n"))
        assert _detect_session_date(rows, []) is None


class TestParseTeamsDatetime:
    def test_none_and_empty(self):
        assert _parse_teams_datetime(None) is None
        assert _parse_teams_datetime("") is None

    @pytest.mark.parametrize("s,expected", [
        ("5/22/26", "2026-05-22"),
        ("5/22/2026", "2026-05-22"),
        ("22-May-2026", "2026-05-22"),
        ("2026-05-22", "2026-05-22"),
    ])
    def test_known_formats(self, s, expected):
        assert _parse_teams_datetime(s) == expected

    def test_takes_date_part_before_comma(self):
        assert _parse_teams_datetime("5/22/26, 10:00:00 AM") == "2026-05-22"

    def test_unparseable_returns_none(self):
        assert _parse_teams_datetime("not a date") is None


class TestNorm:
    def test_lowercases_and_collapses_whitespace(self):
        assert _norm("  John   Doe  ") == "john doe"

    def test_none_returns_empty(self):
        assert _norm(None) == ""


class TestMatchHelpers:
    def test_match_exact(self):
        p_norm = [("john doe", {"raw": "John Doe"})]
        match, i = _match_exact("john doe", p_norm, set())
        assert match["raw"] == "John Doe"
        assert i == 0

    def test_match_exact_skips_used(self):
        p_norm = [("john doe", {"raw": "John Doe"})]
        match, i = _match_exact("john doe", p_norm, {0})
        assert match is None and i is None

    def test_match_partial_all_tokens_present(self):
        p_norm = [("doe john extra", {"raw": "x"})]
        match, i = _match_partial(["john", "doe"], p_norm, set())
        assert match is not None

    def test_match_partial_missing_token_fails(self):
        p_norm = [("john", {"raw": "x"})]
        match, i = _match_partial(["john", "doe"], p_norm, set())
        assert match is None

    def test_match_first_name_empty_parts_returns_none(self):
        assert _match_first_name([], [("x", {})], set()) == (None, None)

    def test_match_first_name(self):
        p_norm = [("john smith", {"raw": "x"})]
        match, i = _match_first_name(["john"], p_norm, set())
        assert match is not None


class TestMatchParticipants:
    def test_exact_match_present(self):
        participants = [{"name": "John Doe", "raw": "John Doe", "minutes": 90.0}]
        results, unmatched = match_participants(["John Doe"], participants, 90.0, 75, 90.0)
        assert results["John Doe"]["present"] is True
        assert results["John Doe"]["matched"] is True
        assert results["John Doe"]["uncertain"] is False
        assert unmatched == []

    def test_below_threshold_marked_absent(self):
        participants = [{"name": "John Doe", "raw": "John Doe", "minutes": 30.0}]
        results, _ = match_participants(["John Doe"], participants, 90.0, 75, 90.0)
        assert results["John Doe"]["present"] is False
        assert results["John Doe"]["matched"] is True

    def test_no_match_marked_not_matched(self):
        results, _ = match_participants(["Ghost"], [], 90.0, 75, 0.0)
        assert results["Ghost"] == {"present": False, "attentiveness": 0.0,
                                    "matched": False, "uncertain": False, "minutes": 0}

    def test_unmatched_participants_listed(self):
        participants = [{"name": "Extra Guest", "raw": "Extra Guest", "minutes": 10.0}]
        _, unmatched = match_participants([], participants, 90.0, 75, 10.0)
        assert unmatched == ["Extra Guest"]

    def test_first_name_only_match_is_uncertain(self):
        participants = [{"name": "John Smith", "raw": "John Smith", "minutes": 90.0}]
        results, _ = match_participants(["John Doe"], participants, 90.0, 75, 90.0)
        assert results["John Doe"]["matched"] is True
        assert results["John Doe"]["uncertain"] is True

    def test_email_match_when_name_shares_no_tokens(self):
        # Teams display name is a personal nickname with zero overlap with the
        # enrolled name — only the email ties the two together.
        participants = [{"name": "rocky", "raw": "Rocky", "minutes": 90.0, "email": "abdul.baseer@corp.com"}]
        results, _ = match_participants(
            ["Abdul Baseer Saleem"], participants, 90.0, 75, 90.0,
            enrolled_emails={"Abdul Baseer Saleem": "Abdul.Baseer@Corp.com"})
        assert results["Abdul Baseer Saleem"]["matched"] is True
        assert results["Abdul Baseer Saleem"]["uncertain"] is False

    def test_email_match_is_case_insensitive(self):
        participants = [{"name": "xyz", "raw": "xyz", "minutes": 90.0, "email": "JOHN@CORP.COM"}]
        results, _ = match_participants(
            ["John Doe"], participants, 90.0, 75, 90.0, enrolled_emails={"John Doe": "john@corp.com"})
        assert results["John Doe"]["matched"] is True

    def test_no_email_match_when_emails_differ(self):
        participants = [{"name": "xyz", "raw": "xyz", "minutes": 90.0, "email": "someone.else@corp.com"}]
        results, _ = match_participants(
            ["John Doe"], participants, 90.0, 75, 90.0, enrolled_emails={"John Doe": "john@corp.com"})
        assert results["John Doe"]["matched"] is False

    def test_fuzzy_match_catches_typo(self):
        # No exact/partial/email match, but the names are a near-identical typo.
        participants = [{"name": "Rahul Sharme", "raw": "Rahul Sharme", "minutes": 90.0}]
        results, _ = match_participants(["Rahul Sharma"], participants, 90.0, 75, 90.0)
        assert results["Rahul Sharma"]["matched"] is True
        assert results["Rahul Sharma"]["uncertain"] is True

    def test_fuzzy_does_not_match_different_people(self):
        # Genuinely different names (no shared token, low character similarity)
        # must not be conflated by any matching tier.
        participants = [{"name": "Kavya Reddy", "raw": "Kavya Reddy", "minutes": 90.0}]
        results, _ = match_participants(["Karthik Iyer"], participants, 90.0, 75, 90.0)
        assert results["Karthik Iyer"]["matched"] is False

    def test_match_fuzzy_helper_below_threshold_returns_none(self):
        match, i = _match_fuzzy("priya singh", [("priya shah", {"raw": "x"})], set())
        assert match is None and i is None

    def test_match_email_helper_no_email_returns_none(self):
        assert _match_email("", {}, set()) == (None, None)

    def test_zero_session_minutes_never_present(self):
        participants = [{"name": "John", "raw": "John", "minutes": 0.0}]
        results, _ = match_participants(["John"], participants, 0.0, 75, 0.0)
        assert results["John"]["present"] is False

    def test_participants_not_double_matched(self):
        participants = [{"name": "John Doe", "raw": "John Doe", "minutes": 90.0}]
        results, unmatched = match_participants(["John Doe", "Johnathan Doeling"], participants, 90.0, 50, 90.0)
        # Only one of the two enrolled names can claim the single participant.
        matched_count = sum(1 for r in results.values() if r["matched"])
        assert matched_count == 1
        assert unmatched == []


class TestAttentivenessRounding:
    def test_attentiveness_rounds_to_whole_percent(self):
        # 67/90 = 0.7444... — must come back as 0.74 (74%), not the old 4-decimal 0.7444.
        participants = [{"name": "John Doe", "raw": "John Doe", "minutes": 67.0}]
        results, _ = match_participants(["John Doe"], participants, 90.0, 50, 90.0)
        assert results["John Doe"]["attentiveness"] == 0.74


class TestSheetHelpers:
    def _tracker(self, names, sub_row=2, name_col=2):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="S.No")
        ws.cell(row=1, column=name_col, value="Name")
        for i, name in enumerate(names, start=sub_row + 1):
            ws.cell(row=i, column=1, value=i - sub_row)
            ws.cell(row=i, column=name_col, value=name)
        return wb, ws

    def test_find_name_column(self):
        wb, ws = self._tracker(["John"])
        assert _find_name_column(ws) == 2

    def test_find_name_column_defaults_to_b(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert _find_name_column(ws) == 2

    def test_find_email_column(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Email")
        assert _find_email_column(ws) == 2

    def test_find_email_column_absent_returns_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")
        assert _find_email_column(ws) is None

    def test_enrolled_rows(self):
        wb, ws = self._tracker(["John", "Jane"])
        rows = _enrolled_rows(ws, 2, 3)
        assert rows == [(3, "John"), (4, "Jane")]

    def test_enrolled_rows_stops_at_blank(self):
        wb, ws = self._tracker(["John"])
        ws.cell(row=5, column=2, value="")
        rows = _enrolled_rows(ws, 2, 3)
        assert rows == [(3, "John")]

    def test_find_subheader_row_majority_vote(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for c in range(1, 4):
            ws.cell(row=2, column=c, value="Attendance")
        ws.cell(row=3, column=1, value="Attentiveness")
        assert _find_subheader_row(ws, 3) == 2

    def test_find_subheader_row_defaults_when_no_match(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert _find_subheader_row(ws, 3) == 2

    def test_find_alert_fill_no_col_returns_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert _find_alert_fill(ws, None, []) == (None, None)

    def test_find_alert_fill_finds_existing_solid_fill(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws.cell(row=3, column=2, value="No")
        cell.fill = PatternFill("solid", fgColor="FFC7CE")
        fill, font_color = _find_alert_fill(ws, 2, [(3, "John")])
        assert fill is not None

    def test_find_alert_fill_no_baked_fill_returns_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=3, column=2, value="No")
        fill, font_color = _find_alert_fill(ws, 2, [(3, "John")])
        assert fill is None

    def test_header_is_merged_true(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        assert _header_is_merged(ws, 1, 2) is True

    def test_header_is_merged_false(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert _header_is_merged(ws, 1, 2) is False


class TestFmtMinutes:
    def test_zero(self):
        assert _fmt_minutes(0) == "0m"

    def test_hours_and_minutes(self):
        assert _fmt_minutes(90) == "1h 30m"

    def test_hours_minutes_seconds(self):
        assert _fmt_minutes(67.5833) == "1h 7m 34s" or _fmt_minutes(67.5833) == "1h 7m 35s"

    def test_minutes_only(self):
        assert _fmt_minutes(45) == "45m"


def _build_tracker_bytes():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Consolidated Report"
    ws1.cell(row=1, column=1, value="S.No")
    ws1.cell(row=1, column=2, value="Name")
    for i, name in enumerate(["Abdul Baseer Saleem", "John Doe", "Jane Roe"], start=3):
        ws1.cell(row=i, column=1, value=i - 2)
        ws1.cell(row=i, column=2, value=name)

    ws2 = wb.create_sheet("Overall Attendance %")
    ws2.cell(row=1, column=1, value="S.No")
    ws2.cell(row=1, column=2, value="Name")
    for i, name in enumerate(["Abdul Baseer Saleem", "John Doe", "Jane Roe"], start=3):
        ws2.cell(row=i, column=1, value=i - 2)
        ws2.cell(row=i, column=2, value=name)

    wb.create_sheet("Login Details")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestProcessAttendance:
    def test_single_session_end_to_end(self):
        tracker_bytes = _build_tracker_bytes()
        output_bytes, output_filename, info = process_attendance(
            tracker_bytes, "Tracker_20thMay.xlsx", _teams_csv(BASIC_EXPORT), "export.csv",
            "Session 1", "2026-05-22", 75,
        )
        assert output_filename == "Tracker_22ndMay.xlsx"
        assert info["enrolled"] == 3
        assert info["present"] == 1
        assert info["absent"] == 2
        assert info["output_filename"] == output_filename

        wb2 = openpyxl.load_workbook(io.BytesIO(output_bytes))
        ws1 = wb2["Consolidated Report"]
        assert ws1.cell(row=3, column=4).value == "Yes"
        assert ws1.cell(row=4, column=4).value == "No"

    def test_batch_multi_session(self):
        tracker_bytes = _build_tracker_bytes()
        second_export = (
            "3. In-Meeting Activities\n"
            "Name,First Join,Last Leave,In-Meeting Duration,Attentiveness\n"
            "Abdul Baseer Saleem,5/23/26 10:00 AM,5/23/26 11:00 AM,1h,\n"
            "John Doe,5/23/26 10:00 AM,5/23/26 11:00 AM,1h,\n"
        )
        sessions = [
            {"bytes": _teams_csv(BASIC_EXPORT), "filename": "e1.csv",
             "session_name": "Session 1", "session_date_iso": "2026-05-22"},
            {"bytes": _teams_csv(second_export), "filename": "e2.csv",
             "session_name": "Session 2", "session_date_iso": "2026-05-23"},
        ]
        out_bytes, out_name, summary = process_attendance_batch(
            tracker_bytes, "Tracker_20thMay.xlsx", sessions, 75)
        assert out_name == "Tracker_23rdMay.xlsx"
        assert summary["sessions_processed"] == 2
        assert len(summary["days"]) == 2
        assert summary["days"][0]["session_name"] == "Session 1"
        assert summary["days"][1]["session_name"] == "Session 2"

    def test_batch_processed_in_chronological_order_regardless_of_input_order(self):
        tracker_bytes = _build_tracker_bytes()
        sessions = [
            {"bytes": _teams_csv(BASIC_EXPORT), "filename": "e2.csv",
             "session_name": "Later", "session_date_iso": "2026-05-23"},
            {"bytes": _teams_csv(BASIC_EXPORT), "filename": "e1.csv",
             "session_name": "Earlier", "session_date_iso": "2026-05-22"},
        ]
        _, _, summary = process_attendance_batch(tracker_bytes, "Tracker.xlsx", sessions, 75)
        assert [d["session_name"] for d in summary["days"]] == ["Earlier", "Later"]

    def test_batch_raises_on_empty_sessions(self):
        with pytest.raises(ValueError):
            process_attendance_batch(_build_tracker_bytes(), "Tracker.xlsx", [], 75)

    def test_avg_attendance_pct_is_a_whole_number(self):
        # 1 present out of 3 enrolled = 33.333...% — must come back as int 33, not 33.3.
        tracker_bytes = _build_tracker_bytes()
        _, _, info = process_attendance(
            tracker_bytes, "Tracker_20thMay.xlsx", _teams_csv(BASIC_EXPORT), "export.csv",
            "Session 1", "2026-05-22", 75,
        )
        assert info["avg_attendance_pct"] == 33
        assert isinstance(info["avg_attendance_pct"], int)

    def test_attentiveness_cell_uses_whole_percent_format(self):
        tracker_bytes = _build_tracker_bytes()
        output_bytes, _, _ = process_attendance(
            tracker_bytes, "Tracker_20thMay.xlsx", _teams_csv(BASIC_EXPORT), "export.csv",
            "Session 1", "2026-05-22", 75,
        )
        wb2 = openpyxl.load_workbook(io.BytesIO(output_bytes))
        ws1 = wb2["Consolidated Report"]
        assert ws1.cell(row=3, column=5).number_format == "0%"

    def test_no_scheduled_minutes_uses_teams_derived_duration(self):
        tracker_bytes = _build_tracker_bytes()
        _, _, info = process_attendance(
            tracker_bytes, "Tracker_20thMay.xlsx", _teams_csv(BASIC_EXPORT), "export.csv",
            "Session 1", "2026-05-22", 75,
        )
        assert info["session_minutes"] == 90.0  # "1h 30m" Meeting Duration, unchanged
        assert info["capped_by_schedule"] is False

    def test_session_overrun_capped_at_scheduled_duration(self):
        # Teams reports a 90-minute meeting; the schedule says only 60 minutes were
        # planned (e.g. the host kept the room open after class actually ended) —
        # attendance/attentiveness must be measured against the scheduled 60, not 90.
        tracker_bytes = _build_tracker_bytes()
        _, _, info = process_attendance(
            tracker_bytes, "Tracker_20thMay.xlsx", _teams_csv(BASIC_EXPORT), "export.csv",
            "Session 1", "2026-05-22", 75, scheduled_minutes=60,
        )
        assert info["session_minutes"] == 60.0
        assert info["capped_by_schedule"] is True

    def test_session_ran_short_uses_actual_not_padded_scheduled_duration(self):
        # Schedule says 120 minutes were planned, but the meeting itself only ran
        # 90 — attendance must be measured against what actually happened (90),
        # not padded out to the longer scheduled time.
        tracker_bytes = _build_tracker_bytes()
        _, _, info = process_attendance(
            tracker_bytes, "Tracker_20thMay.xlsx", _teams_csv(BASIC_EXPORT), "export.csv",
            "Session 1", "2026-05-22", 75, scheduled_minutes=120,
        )
        assert info["session_minutes"] == 90.0
        assert info["capped_by_schedule"] is False

    def test_batch_applies_per_day_scheduled_minutes(self):
        tracker_bytes = _build_tracker_bytes()
        second_export = (
            "3. In-Meeting Activities\n"
            "Name,First Join,Last Leave,In-Meeting Duration,Attentiveness\n"
            "Abdul Baseer Saleem,5/23/26 10:00 AM,5/23/26 11:00 AM,1h,\n"
        )
        sessions = [
            {"bytes": _teams_csv(BASIC_EXPORT), "filename": "e1.csv",
             "session_name": "Session 1", "session_date_iso": "2026-05-22"},
            {"bytes": _teams_csv(second_export), "filename": "e2.csv",
             "session_name": "Session 2", "session_date_iso": "2026-05-23"},
        ]
        _, _, summary = process_attendance_batch(
            tracker_bytes, "Tracker_20thMay.xlsx", sessions, 75,
            scheduled_minutes_by_date={"2026-05-22": 45, "2026-05-23": 90})
        assert summary["days"][0]["session_minutes"] == 45.0  # 90 capped to scheduled 45
        assert summary["days"][0]["capped_by_schedule"] is True
        assert summary["days"][1]["session_minutes"] == 60.0  # 1h actual, under the 90 scheduled
        assert summary["days"][1]["capped_by_schedule"] is False
