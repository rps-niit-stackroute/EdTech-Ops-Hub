"""Integration tests for /api/attendance/*, /api/audit*, /api/admin/backup*, and /api/."""
import io
import json
from urllib.parse import unquote

import openpyxl

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

TEAMS_EXPORT_TEXT = (
    "Summary\n"
    "Meeting Start Time,9/1/26, 10:00:00 AM\n"
    "Meeting Duration,1h\n"
    "\n"
    "3. In-Meeting Activities\n"
    "Name,First Join,Last Leave,In-Meeting Duration,Attentiveness\n"
    "IT_Mentor_A,9/1/26 10:00 AM,9/1/26 11:00 AM,1h,\n"
)


def _teams_bytes():
    return TEAMS_EXPORT_TEXT.encode("utf-16")


def _tracker_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidated Report"
    ws.cell(row=1, column=1, value="S.No")
    ws.cell(row=1, column=2, value="Name")
    ws.cell(row=3, column=1, value=1)
    ws.cell(row=3, column=2, value="IT_Mentor_A")
    wb.create_sheet("Overall Attendance %")
    wb.create_sheet("Login Details")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestAttendanceUpdate:
    def test_single_session_update(self, admin_client, new_program):
        files = {
            "tracker": ("tracker_01stSep.xlsx", _tracker_bytes(), XLSX_CT),
            "teams": ("teams.csv", _teams_bytes(), "text/csv"),
        }
        data = {"session_name": "IT Session", "session_date": "2026-09-01", "threshold": "50",
                "program_id": new_program["id"]}
        r = admin_client.post("/api/attendance/update", files=files, data=data)
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers["content-type"]
        assert "x-process-info" in r.headers
        assert "x-output-filename" in r.headers
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert "Consolidated Report" in wb.sheetnames

    def test_missing_fields_rejected(self, admin_client):
        r = admin_client.post("/api/attendance/update", data={})
        assert r.status_code in (400, 422)

    def test_detect_date_from_teams_export(self, admin_client):
        files = {"teams": ("teams.csv", _teams_bytes(), "text/csv")}
        r = admin_client.post("/api/attendance/detect-date", files=files)
        assert r.status_code == 200
        assert r.json()["session_date"] == "2026-09-01"

    def test_detect_date_unparseable_returns_null(self, admin_client):
        files = {"teams": ("teams.csv", b"garbage,not,a,real,export\n", "text/csv")}
        r = admin_client.post("/api/attendance/detect-date", files=files)
        assert r.status_code == 200
        assert r.json()["session_date"] is None

    def test_session_duration_capped_by_matching_scheduled_session(self, admin_client, new_program):
        # TEAMS_EXPORT_TEXT reports a 1h (60 min) meeting, but the schedule for this
        # exact program+date only planned 30 minutes — the response must show the
        # capped, scheduled duration was actually used, not the longer Teams one.
        admin_client.post("/api/sessions", json={
            "program_id": new_program["id"], "date": "2026-09-01",
            "start_time": "10:00", "end_time": "10:30", "mentor_name": "IT_Mentor_A", "topic": "Capped",
        })
        files = {
            "tracker": ("tracker_01stSep.xlsx", _tracker_bytes(), XLSX_CT),
            "teams": ("teams.csv", _teams_bytes(), "text/csv"),
        }
        data = {"session_name": "IT Session", "session_date": "2026-09-01", "threshold": "50",
                "program_id": new_program["id"]}
        r = admin_client.post("/api/attendance/update", files=files, data=data)
        assert r.status_code == 200, r.text
        info = json.loads(unquote(r.headers["x-process-info"]))
        assert info["session_minutes"] == 30.0
        assert info["capped_by_schedule"] is True

    def test_no_scheduled_session_uses_teams_duration_uncapped(self, admin_client, new_program):
        # No session exists for this program on this date, so it must fall back to
        # the Teams export's own reported duration, unchanged.
        files = {
            "tracker": ("tracker_01stSep.xlsx", _tracker_bytes(), XLSX_CT),
            "teams": ("teams.csv", _teams_bytes(), "text/csv"),
        }
        data = {"session_name": "IT Session", "session_date": "2026-09-01", "threshold": "50",
                "program_id": new_program["id"]}
        r = admin_client.post("/api/attendance/update", files=files, data=data)
        assert r.status_code == 200, r.text
        info = json.loads(unquote(r.headers["x-process-info"]))
        assert info["session_minutes"] == 60.0
        assert info["capped_by_schedule"] is False

    def test_avg_attendance_pct_whole_number_in_response(self, admin_client, new_program):
        files = {
            "tracker": ("tracker_01stSep.xlsx", _tracker_bytes(), XLSX_CT),
            "teams": ("teams.csv", _teams_bytes(), "text/csv"),
        }
        data = {"session_name": "IT Session", "session_date": "2026-09-01", "threshold": "50",
                "program_id": new_program["id"]}
        r = admin_client.post("/api/attendance/update", files=files, data=data)
        info = json.loads(unquote(r.headers["x-process-info"]))
        assert info["avg_attendance_pct"] == int(info["avg_attendance_pct"])


class TestAttendanceBatch:
    def test_batch_update_two_sessions(self, admin_client):
        files = [
            ("teams_files", ("d1.csv", _teams_bytes(), "text/csv")),
            ("teams_files", ("d2.csv", _teams_bytes(), "text/csv")),
        ]
        files.append(("tracker", ("tracker.xlsx", _tracker_bytes(), XLSX_CT)))
        data = {
            "session_names": ["Day 1", "Day 2"],
            "session_dates": ["2026-09-01", "2026-09-02"],
            "threshold": "50",
        }
        r = admin_client.post("/api/attendance/update-batch", files=files, data=data)
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers["content-type"]

    def test_batch_no_teams_files_rejected(self, admin_client):
        files = [("tracker", ("tracker.xlsx", _tracker_bytes(), XLSX_CT))]
        r = admin_client.post("/api/attendance/update-batch", files=files,
                              data={"session_names": [], "session_dates": []})
        assert r.status_code in (400, 422)

    def test_batch_mismatched_lengths_rejected(self, admin_client):
        files = [
            ("teams_files", ("d1.csv", _teams_bytes(), "text/csv")),
            ("tracker", ("tracker.xlsx", _tracker_bytes(), XLSX_CT)),
        ]
        data = {"session_names": ["Only One Name"], "session_dates": ["2026-09-01", "2026-09-02"]}
        r = admin_client.post("/api/attendance/update-batch", files=files, data=data)
        assert r.status_code == 400

    def test_batch_feedback_without_program_id_rejected(self, admin_client):
        files = [
            ("teams_files", ("d1.csv", _teams_bytes(), "text/csv")),
            ("tracker", ("tracker.xlsx", _tracker_bytes(), XLSX_CT)),
            ("feedback", ("feedback.xlsx", _tracker_bytes(), XLSX_CT)),
        ]
        data = {"session_names": ["Day 1"], "session_dates": ["2026-09-01"]}
        r = admin_client.post("/api/attendance/update-batch", files=files, data=data)
        assert r.status_code == 400


class TestAuditLog:
    def test_audit_log_lists_recent_actions(self, admin_client, new_program):
        r = admin_client.get("/api/audit")
        assert r.status_code == 200
        d = r.json()
        assert "rows" in d and "users" in d and "actions" in d
        assert any(row["action"] == "Program created" for row in d["rows"])

    def test_audit_log_filtered_by_action(self, admin_client):
        r = admin_client.get("/api/audit", params={"action": "Program created"})
        assert r.status_code == 200
        assert all(row["action"] == "Program created" for row in r.json()["rows"])

    def test_audit_export_csv(self, admin_client):
        r = admin_client.get("/api/audit/export")
        assert r.status_code == 200
        assert "text/csv" in r.headers["content-type"]
        assert r.content.startswith(b"id,user_name,role,action,details,timestamp")


class TestBackup:
    def test_download_backup_zip(self, admin_client):
        r = admin_client.get("/api/admin/backup")
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/zip"
        assert len(r.content) > 0

    def test_last_backup_meta_after_download(self, admin_client):
        admin_client.get("/api/admin/backup")
        r = admin_client.get("/api/admin/backup/last")
        assert r.status_code == 200
        assert r.json()["last_backup"] is not None
        assert "filename" in r.json()["last_backup"]


class TestRoot:
    def test_root_message(self, client):
        r = client.get("/api/")
        assert r.status_code == 200
        assert r.json()["message"] == "Delivery Automation API"
