"""Backend API tests for EdTech Ops Hub.
Covers: dashboard, programs CRUD, sessions, calendar/clashes, meta, sow, attendance.
"""
import io
import os
from datetime import date

import openpyxl
import pytest
import requests
from dotenv import load_dotenv

load_dotenv("/app/frontend/.env")
BASE = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE}/api"


@pytest.fixture(scope="module")
def s():
    return requests.Session()


# --------------------------- Dashboard ---------------------------
class TestDashboard:
    def test_dashboard_metrics(self, s):
        r = s.get(f"{API}/dashboard", timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        for k in ("total_programs", "sessions_this_week",
                 "active_mentors", "clashes_detected", "total_sessions"):
            assert k in d
        assert d["total_programs"] >= 4
        assert d["clashes_detected"] >= 1
        assert d["active_mentors"] >= 4


# --------------------------- Programs CRUD ---------------------------
class TestPrograms:
    created_id = None

    def test_list_programs(self, s):
        r = s.get(f"{API}/programs", timeout=30)
        assert r.status_code == 200
        progs = r.json()
        assert isinstance(progs, list)
        assert len(progs) >= 4
        sample = progs[0]
        assert "id" in sample
        assert "name" in sample
        assert "session_count" in sample
        assert "mentors" in sample

    def test_create_program(self, s):
        payload = {
            "name": "TEST_Program_QA",
            "client": "TEST_Client",
            "project_code": "TEST-001",
            "team_member": "TEST_Owner",
            "mentors": ["TEST_Mentor"],
        }
        r = s.post(f"{API}/programs", json=payload, timeout=30)
        assert r.status_code == 200, r.text
        p = r.json()
        assert p["name"] == "TEST_Program_QA"
        assert p["client"] == "TEST_Client"
        assert "TEST_Mentor" in p["mentors"]
        assert p["session_count"] == 0
        TestPrograms.created_id = p["id"]

    def test_get_program(self, s):
        assert TestPrograms.created_id
        r = s.get(f"{API}/programs/{TestPrograms.created_id}", timeout=30)
        assert r.status_code == 200
        p = r.json()
        assert p["id"] == TestPrograms.created_id
        assert "sessions" in p and isinstance(p["sessions"], list)

    def test_update_program(self, s):
        assert TestPrograms.created_id
        payload = {
            "name": "TEST_Program_QA_Updated",
            "client": "TEST_Client",
            "project_code": "TEST-001",
            "team_member": "TEST_Owner",
            "mentors": ["TEST_Mentor", "TEST_Mentor2"],
        }
        r = s.put(f"{API}/programs/{TestPrograms.created_id}", json=payload, timeout=30)
        assert r.status_code == 200
        p = r.json()
        assert p["name"] == "TEST_Program_QA_Updated"
        # GET to verify persisted
        r2 = s.get(f"{API}/programs/{TestPrograms.created_id}", timeout=30)
        assert r2.json()["name"] == "TEST_Program_QA_Updated"

    def test_create_session_and_update(self, s):
        assert TestPrograms.created_id
        payload = {
            "program_id": TestPrograms.created_id,
            "date": date.today().strftime("%Y-%m-%d"),
            "start_time": "10:00",
            "end_time": "11:00",
            "topic": "TEST topic",
            "mentor_name": "TEST_Mentor",
        }
        r = s.post(f"{API}/sessions", json=payload, timeout=30)
        assert r.status_code == 200, r.text
        sid = r.json()["id"]
        # update
        r2 = s.put(f"{API}/sessions/{sid}",
                   json={"start_time": "11:00", "end_time": "12:30"}, timeout=30)
        assert r2.status_code == 200
        assert r2.json()["start_time"] == "11:00"
        # delete
        r3 = s.delete(f"{API}/sessions/{sid}", timeout=30)
        assert r3.status_code == 200

    def test_delete_program(self, s):
        assert TestPrograms.created_id
        r = s.delete(f"{API}/programs/{TestPrograms.created_id}", timeout=30)
        assert r.status_code == 200
        r2 = s.get(f"{API}/programs/{TestPrograms.created_id}", timeout=30)
        assert r2.status_code == 404


# --------------------------- Meta ---------------------------
class TestMeta:
    def test_meta(self, s):
        r = s.get(f"{API}/meta", timeout=30)
        assert r.status_code == 200
        d = r.json()
        for k in ("mentors", "team_members", "programs"):
            assert k in d
        assert len(d["mentors"]) >= 4
        assert "Manoj Ajmer" in d["mentors"]


# --------------------------- Calendar / Clashes ---------------------------
class TestCalendar:
    def test_calendar_current_month(self, s):
        today = date.today()
        r = s.get(f"{API}/calendar", params={"month": today.month, "year": today.year}, timeout=30)
        assert r.status_code == 200
        d = r.json()
        assert "sessions" in d and "clashes" in d
        # has_clash flag set on at least one session
        clashed = [x for x in d["sessions"] if x.get("has_clash")]
        assert len(clashed) >= 2, "Expected at least 2 sessions flagged with has_clash"

    def test_clashes_endpoint(self, s):
        r = s.get(f"{API}/clashes", timeout=30)
        assert r.status_code == 200
        cl = r.json()["clashes"]
        assert len(cl) >= 1
        c0 = cl[0]
        for k in ("mentor", "date", "program_a", "program_b", "time_a", "time_b"):
            assert k in c0
        assert c0["mentor"].lower().startswith("manoj")


# --------------------------- SOW ---------------------------
class TestSOW:
    def test_sow_preview(self, s):
        today = date.today()
        r = s.get(f"{API}/sow", params={"month": today.month, "year": today.year}, timeout=30)
        assert r.status_code == 200
        d = r.json()
        assert "grouped" in d and "grand_total" in d and "month_label" in d
        assert d["grand_total"]["sessions"] >= 1
        # subtotal structure
        if d["grouped"]:
            g = d["grouped"][0]
            for k in ("mentor", "rows", "subtotal_sessions", "subtotal_hours"):
                assert k in g

    def test_sow_download(self, s):
        today = date.today()
        r = s.get(f"{API}/sow/download",
                  params={"month": today.month, "year": today.year}, timeout=30)
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct, ct
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert len(wb.sheetnames) >= 1


# --------------------------- Attendance ---------------------------
def _make_teams_csv():
    """Synthesize a tab-separated Teams export with header sections."""
    lines = [
        "1. Summary",
        "Meeting Title\tTest Meeting",
        "Meeting Duration\t1h 30m",
        "",
        "2. Participants",
        "Name\tFirst Join\tLast Leave\tIn-Meeting Duration\tAttentiveness",
        "Manoj Ajmer\t9:30 AM\t12:00 PM\t1h 30m\t0.95",
        "Kishore K\t9:35 AM\t12:00 PM\t1h 25m\t0.90",
        "Unknown Person\t9:40 AM\t12:00 PM\t1h 20m\t0.85",
        "",
        "3. In-Meeting Activities",
    ]
    return "\n".join(lines).encode("utf-16")


class TestAttendance:
    def test_attendance_update(self, s):
        tracker_path = "/app/samples/tracker.xlsx"
        if not os.path.exists(tracker_path):
            pytest.skip("tracker.xlsx sample not available")
        with open(tracker_path, "rb") as f:
            tracker_bytes = f.read()
        teams_bytes = _make_teams_csv()
        files = {
            "tracker": ("tracker_22-05-26.xlsx", tracker_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            "teams": ("teams_export.csv", teams_bytes, "text/csv"),
        }
        data = {
            "session_name": "TEST_Session",
            "session_date": date.today().strftime("%Y-%m-%d"),
            "threshold": 50,
        }
        r = s.post(f"{API}/attendance/update", files=files, data=data, timeout=60)
        assert r.status_code == 200, r.text[:500]
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct
        assert "X-Process-Info" in r.headers
        assert "X-Output-Filename" in r.headers
        # Output xlsx must open
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert len(wb.sheetnames) >= 1

    def test_attendance_missing_fields(self, s):
        r = s.post(f"{API}/attendance/update", data={}, timeout=30)
        assert r.status_code in (400, 422)
