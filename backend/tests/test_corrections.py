"""Tests for the 6 corrections (iteration 2):
1) Attentiveness computed = dur / max_dur, formatted 0.00%, absent=0
2) Appended cells carry styling (fill + borders + font)
3) /api/sow returns 'dates' per row
4) /api/sow/download has NO per-mentor subtotal rows, only data + grand total
5) S.No first column in preview + download; GRAND TOTAL S.No blank
6) /api/attendance/detect-date returns {session_date}
Regression handled in test_backend.py.
"""
import io
import os
import re
from datetime import date

import openpyxl
import pytest
import requests
from dotenv import load_dotenv

load_dotenv("/app/frontend/.env")
BASE = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE}/api"

TRACKER = "/app/samples/tracker.xlsx"


def _make_teams_csv(meeting_start="6/15/26, 9:30:00 AM", duration="1h 30m"):
    """Synthesize a UTF-16 tab-separated Teams export."""
    lines = [
        "1. Summary",
        "Meeting Title\tTest Meeting",
        f"Meeting Start Time\t{meeting_start}",
        f"Meeting Duration\t{duration}",
        "",
        "2. Participants",
        "Name\tFirst Join\tLast Leave\tIn-Meeting Duration",
        # Highest = 2h 28m = 148 min  -> should map to 100% (enrolled in tracker)
        "Seshadri Rajan\t9:30 AM\t11:58 AM\t2h 28m",
        # 1h 5m  = 65 min  -> 65/148 = ~43.92% (enrolled)
        "Mrudul Gupta\t9:35 AM\t10:40 AM\t1h 5m",
        "Unknown Person\t9:40 AM\t11:00 AM\t1h 20m",
        "",
        "3. In-Meeting Activities",
    ]
    return "\n".join(lines).encode("utf-16")


@pytest.fixture(scope="module")
def s():
    return requests.Session()


# ===== CORRECTION 6 — Auto-detect session date =====
class TestDetectDate:
    def test_detect_date_returns_iso(self, s):
        files = {"teams": ("teams.csv", _make_teams_csv("6/15/26, 9:30:00 AM"), "text/csv")}
        r = s.post(f"{API}/attendance/detect-date", files=files, timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d.get("session_date") == "2026-06-15", d

    def test_detect_date_missing_file_422(self, s):
        r = s.post(f"{API}/attendance/detect-date", timeout=15)
        assert r.status_code == 422


# ===== CORRECTION 1 & 2 — Attentiveness computed + styling =====
class TestAttendanceCorrections:
    @pytest.fixture(scope="class")
    def output_wb(self, s):
        if not os.path.exists(TRACKER):
            pytest.skip("tracker.xlsx not available")
        with open(TRACKER, "rb") as f:
            tracker_bytes = f.read()
        teams_bytes = _make_teams_csv()
        files = {
            "tracker": ("tracker_22-05-26.xlsx", tracker_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            "teams": ("teams_export.csv", teams_bytes, "text/csv"),
        }
        data = {"session_name": "TEST_QA_Session",
                "session_date": "2026-06-15", "threshold": 50}
        r = s.post(f"{API}/attendance/update", files=files, data=data, timeout=60)
        assert r.status_code == 200, r.text[:500]
        return openpyxl.load_workbook(io.BytesIO(r.content))

    def _find_sheet(self, wb, keys):
        for sn in wb.sheetnames:
            low = sn.lower()
            if any(k in low for k in keys):
                return wb[sn]
        return None

    def test_consolidated_attentiveness_values_and_format(self, output_wb):
        ws = self._find_sheet(output_wb, ["consolidated", "report"])
        assert ws is not None
        max_col = ws.max_column
        # The last 2 appended columns are Attendance/Attentiveness
        att_col = max_col - 1
        atten_col = max_col
        # Row 2 should have Attendance / Attentiveness labels
        assert str(ws.cell(row=2, column=att_col).value).strip().lower() == "attendance"
        assert str(ws.cell(row=2, column=atten_col).value).strip().lower() == "attentiveness"

        # Scan data rows; verify number_format 0.00% and presence of 1.0 (100%) somewhere
        found_full = False
        found_partial = False
        absent_zero = False
        for r in range(3, ws.max_row + 1):
            name = ws.cell(row=r, column=2).value
            if not name:
                continue
            v = ws.cell(row=r, column=atten_col).value
            fmt = ws.cell(row=r, column=atten_col).number_format
            assert fmt == "0.00%", f"Row {r} fmt={fmt}"
            if isinstance(v, (int, float)):
                if abs(v - 1.0) < 1e-3:
                    found_full = True
                if 0.0 < v < 1.0:
                    found_partial = True
                if v == 0:
                    absent_zero = True
        assert found_full, "Expected a 100% attentiveness for highest-duration matched participant"
        # absent enrolled participants should be 0
        assert absent_zero, "Expected at least one absent participant attentiveness = 0"
        # Optional: a partial value confirms ratio works
        assert found_partial or found_full

    def test_consolidated_appended_cells_styled(self, output_wb):
        ws = self._find_sheet(output_wb, ["consolidated", "report"])
        max_col = ws.max_column
        # Row1 merged header + Row2 sub-headers must have solid fill + border + bold-ish font
        for (r, c) in [(1, max_col - 1), (2, max_col - 1), (2, max_col)]:
            cell = ws.cell(row=r, column=c)
            assert cell.fill is not None and cell.fill.patternType == "solid", \
                f"cell ({r},{c}) fill not solid: {cell.fill.patternType}"
            b = cell.border
            assert any([b.left and b.left.style, b.right and b.right.style,
                        b.top and b.top.style, b.bottom and b.bottom.style]), \
                f"cell ({r},{c}) has no border"
            assert cell.font is not None
        # Data rows: borders copied
        cell = ws.cell(row=3, column=max_col)
        b = cell.border
        assert any([b.left and b.left.style, b.right and b.right.style,
                    b.top and b.top.style, b.bottom and b.bottom.style]), \
            "data row has no border"

    def test_login_sheet_appended_styled(self, output_wb):
        ws = self._find_sheet(output_wb, ["login"])
        if ws is None:
            pytest.skip("Login sheet not present")
        # Find rightmost 'Name' subheader in row 3 -> new group start
        starts = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=3, column=c).value
            if v and str(v).strip().lower() == "name":
                starts.append(c)
        assert len(starts) >= 1
        new_start = starts[-1]
        # Row1 merged header solid fill
        hdr = ws.cell(row=1, column=new_start)
        assert hdr.fill is not None and hdr.fill.patternType == "solid", \
            f"login header fill not solid: {hdr.fill.patternType}"
        # Sub-headers row3 (Name, First Join, Last Leave, In-Meeting Duration)
        for off in range(4):
            sub = ws.cell(row=3, column=new_start + off)
            assert sub.fill is not None and sub.fill.patternType == "solid", \
                f"login sub-header ({new_start + off}) fill not solid"


# ===== CORRECTION 3 — SOW dates field =====
class TestSOWDates:
    def test_sow_preview_has_dates(self, s):
        r = s.get(f"{API}/sow", params={"month": 6, "year": 2026}, timeout=30)
        assert r.status_code == 200
        d = r.json()
        # Find a row and verify dates field
        any_row = None
        for g in d["grouped"]:
            if g["rows"]:
                any_row = g["rows"][0]
                break
        assert any_row is not None, "No SOW rows for 6/2026"
        assert "dates" in any_row
        # Format like '5 Jun, 8 Jun' — has space + Jun, comma-separated
        assert re.match(r"^\d{1,2}\s+\w{3}(,\s+\d{1,2}\s+\w{3})*$", any_row["dates"]), any_row["dates"]


# ===== CORRECTIONS 4 & 5 — SOW Excel: S.No first, no subtotals, grand total S.No blank =====
class TestSOWExcel:
    @pytest.fixture(scope="class")
    def wb(self, s):
        r = s.get(f"{API}/sow/download", params={"month": 6, "year": 2026}, timeout=30)
        assert r.status_code == 200
        return openpyxl.load_workbook(io.BytesIO(r.content))

    def test_headers_have_sno_first_and_dates_last(self, wb):
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers[0] == "S.No", headers
        assert "Dates" in headers, headers
        # Total Hours should come before Dates
        assert headers.index("Total Hours") < headers.index("Dates")

    def test_no_subtotal_rows_only_grand_total(self, wb):
        ws = wb.active
        col_b_vals = [str(ws.cell(row=r, column=2).value or "") for r in range(2, ws.max_row + 1)]
        # No row should contain "Subtotal"
        for v in col_b_vals:
            assert "subtotal" not in v.lower(), f"Found subtotal-like row: {v}"
        # Last row must be GRAND TOTAL
        assert col_b_vals[-1].strip().upper() == "GRAND TOTAL", col_b_vals

    def test_grand_total_sno_blank_and_sno_sequential(self, wb):
        ws = wb.active
        last_row = ws.max_row
        assert ws.cell(row=last_row, column=1).value in (None, ""), \
            f"Grand total S.No must be blank, got {ws.cell(row=last_row, column=1).value!r}"
        # Data rows sequential 1..N
        expected = 1
        for r in range(2, last_row):
            sno = ws.cell(row=r, column=1).value
            # All data rows must have an int S.No (no subtotals allowed)
            assert isinstance(sno, int), f"Row {r} S.No should be int, got {sno!r}"
            assert sno == expected, f"Row {r}: expected S.No {expected}, got {sno}"
            expected += 1

    def test_dates_cell_format(self, wb):
        ws = wb.active
        # Find Dates column index
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        dates_col = headers.index("Dates") + 1
        # At least one data row should have a dates string
        for r in range(2, ws.max_row):  # skip grand total
            v = ws.cell(row=r, column=dates_col).value
            if v:
                assert re.match(r"^\d{1,2}\s+\w{3}", str(v)), v
                return
        pytest.fail("No dates value found in any data row")
