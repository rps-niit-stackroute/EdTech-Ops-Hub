"""Consolidate a raw feedback-form export (Microsoft Forms style: one row per
submission, no session/date/mentor columns) into a tracker's existing Feedback
sheet — filling in Date/Module Name/Faculty from the program's schedule and
skipping any date already captured, so re-running with an overlapping export
doesn't duplicate rows."""
import io
from datetime import datetime

import openpyxl

from excel_utils import copy_cell_style, sanitize_excel_value


def _norm(s):
    return " ".join(str(s or "").strip().lower().split())


def _find_col(header_map, *substrings):
    for name, idx in header_map.items():
        if all(sub in name for sub in substrings):
            return idx
    return None


def _header_map(ws, row=1):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=c).value
        if v:
            m[_norm(v)] = c
    return m


def _as_datetime(value):
    if isinstance(value, datetime):
        return value
    return None


def parse_feedback_export(file_bytes):
    """Returns a list of {dt, participant_name, takeaways, rating, specific_feedback,
    other_feedback} parsed from a raw feedback-form export (one row per submission)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    header_map = _header_map(ws)

    col_participant = _find_col(header_map, "participant", "name")
    col_completion = _find_col(header_map, "completion", "time")
    col_start = _find_col(header_map, "start", "time")
    col_takeaways = _find_col(header_map, "takeaway")
    col_rating = _find_col(header_map, "scale") or _find_col(header_map, "rating") \
        or _find_col(header_map, "effective")
    col_specific = _find_col(header_map, "specific", "feedback")
    col_other = _find_col(header_map, "any", "other", "feedback")

    rows = []
    for r in range(2, ws.max_row + 1):
        def cell(col):
            return ws.cell(row=r, column=col).value if col else None

        participant = cell(col_participant)
        if not participant or not str(participant).strip():
            continue
        dt = _as_datetime(cell(col_completion)) or _as_datetime(cell(col_start))
        rows.append({
            "dt": dt,
            "participant_name": str(participant).strip(),
            "takeaways": cell(col_takeaways),
            "rating": cell(col_rating),
            "specific_feedback": cell(col_specific),
            "other_feedback": cell(col_other),
        })
    return rows


def _find_feedback_sheet(wb):
    for sn in wb.sheetnames:
        if "feedback" in sn.lower():
            return wb[sn]
    return None


def _feedback_columns(header_map):
    return {
        "sno": _find_col(header_map, "sno") or 1,
        "date": _find_col(header_map, "date"),
        "module": _find_col(header_map, "module"),
        "faculty": _find_col(header_map, "faculty"),
        "participant": _find_col(header_map, "participant", "name"),
        "takeaways": _find_col(header_map, "takeaway"),
        "rating": _find_col(header_map, "rating"),
        "specific": _find_col(header_map, "specific", "feedback"),
        "other": _find_col(header_map, "any", "other", "feedback"),
    }


def _scan_existing_feedback_rows(ws, cols):
    """(existing_dates, max_sno, last_data_row) already captured in the sheet."""
    existing_dates = set()
    max_sno = 0
    last_data_row = 1
    for r in range(2, ws.max_row + 1):
        sno_val = ws.cell(row=r, column=cols["sno"]).value
        date_val = ws.cell(row=r, column=cols["date"]).value
        if sno_val is None and date_val is None:
            continue
        last_data_row = r
        if isinstance(sno_val, (int, float)):
            max_sno = max(max_sno, int(sno_val))
        d = _as_datetime(date_val)
        if d:
            existing_dates.add(d.date().isoformat())
    return existing_dates, max_sno, last_data_row


def _row_values(row, session, cols, next_sno):
    return {
        cols["sno"]: next_sno,
        cols["date"]: row["dt"],
        cols["module"]: session.get("module_name", ""),
        cols["faculty"]: session.get("faculty", ""),
        cols["participant"]: sanitize_excel_value(row["participant_name"]),
        cols["takeaways"]: sanitize_excel_value(row["takeaways"]),
        cols["rating"]: row["rating"],
        cols["specific"]: sanitize_excel_value(row["specific_feedback"]),
        cols["other"]: sanitize_excel_value(row["other_feedback"]),
    }


def _write_feedback_row(ws, row, session, cols, next_sno, next_row, template_row):
    for col, val in _row_values(row, session, cols, next_sno).items():
        if col is None:
            continue
        cell = ws.cell(row=next_row, column=col, value=val)
        if template_row:
            copy_cell_style(ws.cell(row=template_row, column=col), cell)


def _process_feedback_row(ws, row, cols, sessions_by_date, existing_dates, template_row, next_sno, next_row):
    """Writes one feedback row if matched. Returns ("added" | "skipped_existing" | "unmatched", iso_date)."""
    iso_date = row["dt"].date().isoformat()
    if iso_date in existing_dates:
        return "skipped_existing", iso_date
    session = sessions_by_date.get(iso_date)
    if not session:
        return "unmatched", iso_date
    _write_feedback_row(ws, row, session, cols, next_sno, next_row, template_row)
    return "added", iso_date


def append_feedback_sheet(wb, feedback_rows, sessions_by_date):
    """Mutates wb's Feedback sheet in place. sessions_by_date: {iso_date: {module_name,
    faculty}}. Returns info dict summarizing what was added/skipped."""
    empty_result = {"added": 0, "skipped_existing_dates": 0, "unmatched_rows": 0, "unmatched_dates": []}
    ws = _find_feedback_sheet(wb)
    if ws is None:
        return {"sheet_found": False, **empty_result}

    cols = _feedback_columns(_header_map(ws))
    if cols["date"] is None or cols["participant"] is None:
        return {"sheet_found": True, "columns_missing": True, **empty_result}

    existing_dates, max_sno, last_data_row = _scan_existing_feedback_rows(ws, cols)
    template_row = last_data_row if last_data_row > 1 else None
    next_sno = max_sno + 1
    next_row = last_data_row + 1

    added = 0
    skipped_existing = 0
    unmatched_rows = 0
    unmatched_dates = set()

    dated_rows = sorted((r for r in feedback_rows if r["dt"] is not None), key=lambda r: r["dt"])
    for row in dated_rows:
        status, iso_date = _process_feedback_row(
            ws, row, cols, sessions_by_date, existing_dates, template_row, next_sno, next_row)
        if status == "skipped_existing":
            skipped_existing += 1
        elif status == "unmatched":
            unmatched_rows += 1
            unmatched_dates.add(iso_date)
        else:
            next_sno += 1
            next_row += 1
            added += 1

    return {
        "sheet_found": True,
        "added": added,
        "skipped_existing_dates": skipped_existing,
        "unmatched_rows": unmatched_rows,
        "unmatched_dates": sorted(unmatched_dates),
    }
