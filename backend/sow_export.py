"""Generate a styled SOW Excel from aggregated rows."""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from excel_utils import sanitize_excel_value


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
GRAND_FILL = PatternFill("solid", fgColor="FDE68A")
CHANGED_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="9CA3AF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = ["Month", "Trainer", "Duration Hours", "Training", "Training Start date of Sow",
           "Training End date of Sow", "Customer", "Project Code", "Project Manager",
           "Session Dates", "Session Count", "Remarks"]


def _normalize_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(round(value, 2)).rstrip("0").rstrip(".")
    return str(value).strip()


def _format_value(value):
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(round(value, 2)).rstrip("0").rstrip(".")
    return str(value)


def _load_previous_cells(previous_sow_bytes):
    if not previous_sow_bytes:
        return {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(previous_sow_bytes), data_only=True)
        ws = wb.active
        return {(cell.row, cell.column): cell.value for row in ws.iter_rows() for cell in row}
    except Exception:
        return {}


def _write_header(ws, columns):
    for ci, col in enumerate(columns, start=1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")


def _write_row_with_diff(ws, r, vals, previous_cells, previous_sow_bytes):
    """Writes one row, highlighting any cell that changed from previous_cells (the
    last-downloaded SOW) so schedule drift since then is visible at a glance."""
    for ci, v in enumerate(vals, start=1):
        old_value = previous_cells.get((r, ci)) if previous_sow_bytes else None
        if previous_sow_bytes and _normalize_cell_value(old_value) != _normalize_cell_value(v):
            cell = ws.cell(row=r, column=ci,
                           value=f"Old: {_format_value(old_value)} → New: {_format_value(v)}")
            cell.fill = CHANGED_FILL
            cell.font = Font(bold=True)
        else:
            cell = ws.cell(row=r, column=ci, value=v)
        cell.border = BORDER


def build_sow_excel(grouped, month_label, previous_sow_bytes=None):
    """grouped: list of {mentor, rows:[...]}. No subtotal rows; only a grand total at the bottom."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"SOW - {month_label}"[:31]
    _write_header(ws, COLUMNS)

    previous_cells = _load_previous_cells(previous_sow_bytes)
    r = 2
    grand_sessions = 0
    grand_hours = 0.0
    for grp in grouped:
        for row in grp["rows"]:
            vals = [row["month"], sanitize_excel_value(row["mentor"]), row["total_hours"],
                    sanitize_excel_value(row["program_name"]), row.get("start_date", ""),
                    row.get("end_date", ""), sanitize_excel_value(row["client"]),
                    sanitize_excel_value(row["project_code"]), row.get("project_manager", "Santosh"),
                    row.get("dates", ""), row["sessions_conducted"], ""]
            _write_row_with_diff(ws, r, vals, previous_cells, previous_sow_bytes)
            grand_sessions += row["sessions_conducted"]
            grand_hours += row["total_hours"]
            r += 1

    # grand total
    ws.cell(row=r, column=2, value="GRAND TOTAL")
    for ci in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=r, column=ci)
        cell.fill = GRAND_FILL
        cell.font = Font(bold=True)
        cell.border = BORDER
    total_hours_cell = ws.cell(row=r, column=3, value=round(grand_hours, 2))
    total_sessions_cell = ws.cell(row=r, column=11, value=grand_sessions)
    if previous_sow_bytes:
        old_hours = previous_cells.get((r, 3))
        old_sessions = previous_cells.get((r, 11))
        if _normalize_cell_value(old_hours) != _normalize_cell_value(round(grand_hours, 2)):
            total_hours_cell.value = f"Old: {_format_value(old_hours)} → New: {_format_value(round(grand_hours, 2))}"
            total_hours_cell.fill = CHANGED_FILL
            total_hours_cell.font = Font(bold=True)
        if _normalize_cell_value(old_sessions) != _normalize_cell_value(grand_sessions):
            total_sessions_cell.value = f"Old: {_format_value(old_sessions)} → New: {_format_value(grand_sessions)}"
            total_sessions_cell.fill = CHANGED_FILL
            total_sessions_cell.font = Font(bold=True)

    widths = [10, 20, 14, 34, 18, 18, 14, 26, 16, 30, 14, 24]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


PROVISION_COLUMNS = ["Month", "Trainer", "Duration Hours", "Cost per hr", "Total Cost", "Training",
                     "Training Start date of Sow", "Training End date of Sow", "Customer",
                     "Project Code", "Project Manager", "Session Dates", "Session Count"]


def build_provision_excel(rows, charges, month_label):
    """rows: session-based Provision entries (mentor, hours, cost/hr, total cost, ...).
    charges: flat ad-hoc service-charge rows (trainer, description, total_cost only)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Provision - {month_label}"[:31]
    _write_header(ws, PROVISION_COLUMNS)

    r = 2
    grand_hours = 0.0
    grand_cost = 0.0
    for row in rows:
        vals = [row["month"], sanitize_excel_value(row["mentor"]), row["total_hours"],
                row["cost_per_hour"], row["total_cost"], sanitize_excel_value(row["program_name"]),
                row.get("start_date", ""), row.get("end_date", ""), sanitize_excel_value(row["client"]),
                sanitize_excel_value(row["project_code"]), row.get("project_manager", "Santosh"),
                row.get("dates", ""), row["sessions_conducted"]]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = BORDER
        grand_hours += row["total_hours"]
        grand_cost += row["total_cost"]
        r += 1

    for ch in charges:
        vals = [ch["month_label"], sanitize_excel_value(ch["trainer"]), "NA", "NA", ch["total_cost"],
                sanitize_excel_value(ch["description"]), "", "", "", "", "Santosh", "", ""]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = BORDER
        grand_cost += ch["total_cost"]
        r += 1

    ws.cell(row=r, column=2, value="GRAND TOTAL")
    for ci in range(1, len(PROVISION_COLUMNS) + 1):
        cell = ws.cell(row=r, column=ci)
        cell.fill = GRAND_FILL
        cell.font = Font(bold=True)
        cell.border = BORDER
    ws.cell(row=r, column=3, value=round(grand_hours, 2))
    ws.cell(row=r, column=5, value=round(grand_cost, 2))

    widths = [10, 18, 14, 12, 14, 34, 18, 18, 14, 26, 16, 30, 14]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
